#!/usr/bin/env python3
"""
load_to_mdb.py

Import inscriptions from a registration workbook (xlsx "Attendees"
sheet) into an existing SPLASH Meet Manager 11 meet database.

The supplied .mdb is the **authoritative event template**: the script
never creates SWIMSTYLE / SWIMEVENT / AGEGROUP / SWIMSESSION or
COMBINEDEVENTS rows.  It only inserts CLUB / ATHLETE / SWIMRESULT /
RELAY / RELAYPOSITION.  The meet organiser sets up the event
structure (events, age groups, sessions, cumulatifs) in SPLASH; we
just populate entries.

First run vs re-run is auto-detected from whether any SWIMRESULT
rows already exist in the supplied .mdb.  A re-run is additive:
only missing rows are inserted; entry times are updated only when
a faster time is supplied.

Every xlsx ticket must resolve to an existing SWIMEVENT + AGEGROUP
in the template.  Any mismatch is reported as a FATAL error and no
writes are performed.

Age-bracket routing:
  - "15-18" ticket    → AGEGROUP [15, 18]
  - "Open" ticket     → AGEGROUP [19, 99]
  - "MA" individual   → 5-year Masters bracket containing athlete age
  - "MA Relais Mixte" → age-sum bracket containing the squad's total age

Usage:
    python3 load_to_mdb.py --xlsx inscriptions.xlsx --mdb template.mdb
                           [--dry-run]
"""
from __future__ import annotations

import argparse
import datetime as dt
import glob
import os
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import jaydebeapi  # JDBC bridge over UCanAccess

# --------------------------------------------------------------------------- #
# Configuration
# --------------------------------------------------------------------------- #
UCANACCESS_DIR = os.environ.get(
    "UCANACCESS_DIR", "/tmp/ucanaccess/UCanAccess-5.0.1.bin"
)

MEET_NATION = "CAN"
# Age reference date used when routing Masters entries to 5-year brackets
# and when computing the sum-of-ages for Masters relays.  The
# Age date — read at runtime from BSGLOBAL.MEETVALUES.AGEDATE in the
# template MDB.  Initialized to None; TemplateIndex.__init__ sets it.
AGE_DATE: dt.date | None = None

# SPLASH/Lenex gender encoding in SMALLINT columns.
GENDER_MALE   = 1
GENDER_FEMALE = 2
GENDER_ALL    = 0
GENDER_MIXED  = 3     # template uses 3 for mixed relays (check on load)

# SPLASH ROUND values found in the template.  2 = Prelim, 9 = Final,
# 1 = Timed Final (no advancement).
ROUND_TIMED_FINAL = 1
ROUND_PRELIM      = 2
ROUND_FINAL       = 9

# ----------------------------------------------------------------------------
# Ticket-type parsing
# ----------------------------------------------------------------------------
NON_RACE_PREFIXES = (
    "Banquet", "Coach", "Cosmod", "Couloir", "Officiel", "Priorit",
    "Sheraton",
)

# Each xlsx ticket label maps to a catalog UNIQUEID (which must exist in
# the template .mdb) and a flag saying whether it's an individual or
# relay entry.  UIDs below match the Championnats canadiens 2026 template
# ("Championnats canadiens Québec 29-31 mai 2026.mdb"):
#
#   UID 501  200 m Nage avec obstacles / Obstacle Swim         (15-18, Open)
#   UID 541  100 m Nage avec obstacles / Obstacle Swim         (Masters)
#   UID 502  100 m Portage Mannequin palmes / Manikin Carry Fins
#   UID 507  50 m Portage du mannequin plein / Manikin Carry
#   UID 506  100 m Remorquage mannequin palmes / Manikin Tow Fins
#   UID 508  200 m Sauveteur d'acier / Super Lifesaver
#   UID 531  100 m Sauvetage combiné / Rescue Medley
#   UID 504  12 m Lancer de la corde / Line Throw
#   UID 542  4 x 50 m Relais obstacle mixte / Mixed Obstacle Relay
#   UID 543  2 x 50 m  Relais mixte portage / Mixed Carry Relay
#   UID 544  4 x 50 m Relais mixte sauve combiné / Mixed Medley Relay
#
# Masters-only Obstacle maps to UID 541 (100 m); everyone else maps to UID 501
# (200 m).  The parser returns UID 541 when the ticket is "MA <F/M> Obstacle
# 100 m" and UID 501 otherwise.
#
# key    : (ticket_label, is_relay, is_masters_obstacle)
# value  : UNIQUEID
TICKET_UID: dict[tuple, int] = {
    ("Corde",             True,  False): 504,    # 12 m duo relay (RELAYCOUNT=2)
    ("Obstacle",          False, False): 501,    # 15-18 / Open
    ("Obstacle",          False, True):  541,    # Masters 100 m
    ("Portage",           False, False): 502,    # 100 m
    ("Portage50",         False, False): 507,    # 50 m
    ("Remorquage",        False, False): 506,
    ("Sauveteur d'acier", False, False): 508,
    ("Medley",            False, False): 531,    # Sauvetage combiné 100 m

    ("Medley",            True,  False): 544,    # 4 x 50 m Mixed Medley Relay
    ("Obstacle",          True,  False): 542,    # 4 x 50 m Mixed Obstacle Relay
    ("Portage",           True,  False): 543,    # 2 x 50 m Mixed Carry Relay
}

# Age bracket codes and the nominal (AGEMIN, AGEMAX) we're looking for in
# the template for each.  "MASTERS" doesn't map to a single pair — the
# loader picks the 5-year bracket containing the athlete's age at AGE_DATE.
AGE_GROUPS = {  # code -> (AGEMIN, AGEMAX, display name)
    "1518":    (15, 18, "15-18 ans"),
    "OPEN":    (19, 99, "Open (19 & over)"),
    "MASTERS": (None, None, "Maîtres (5-year brackets)"),
}


@dataclass
class EventKey:
    """Uniquely identifies a ticket class.  All attributes come from
    the xlsx; the matching SWIMEVENT / AGEGROUP in the template .mdb
    is resolved at validate/insert time."""
    age_code: str          # '1518' | 'MASTERS' | 'OPEN'
    gender: int            # GENDER_MALE | GENDER_FEMALE | GENDER_MIXED
    uniqueid: int          # template SWIMSTYLE.UNIQUEID
    is_relay: bool

    def key(self) -> tuple:
        return (self.age_code, self.gender, self.uniqueid, self.is_relay)

    @property
    def label(self) -> str:
        g = {1: "M", 2: "F", 3: "X"}.get(self.gender, "?")
        kind = "relay" if self.is_relay else "ind"
        return f"{self.age_code}/{g}/UID{self.uniqueid}/{kind}"


def parse_ticket(ticket: str) -> EventKey | None:
    """Return an EventKey if the ticket is a race, else None."""
    t = ticket.strip()
    for p in NON_RACE_PREFIXES:
        if t.startswith(p):
            return None
    m = re.match(r"^(15-18|MA|Open)\s+(.*)$", t)
    if not m:
        return None
    age_code = {"15-18": "1518", "MA": "MASTERS", "Open": "OPEN"}[m.group(1)]
    rest = m.group(2).strip()

    # Relay?
    mr = re.match(r"^Relais Mixte\s+(\S+)", rest)
    if mr:
        style = mr.group(1).strip()   # 'Medley' | 'Obstacle' | 'Portage'
        uid = TICKET_UID.get((style, True, False))
        if uid is None:
            return None
        return EventKey(age_code, GENDER_MIXED, uid, is_relay=True)

    # Individual: "<F|M> <label> [<n> m]"
    mi = re.match(r"^([FM])\s+(.*)$", rest)
    if not mi:
        return None
    gender = GENDER_MALE if mi.group(1) == "M" else GENDER_FEMALE
    body = mi.group(2).strip()
    mb = re.match(r"^(.*?)(?:\s+(\d+)\s*m)?$", body)
    label = mb.group(1).strip()
    dist_txt = mb.group(2)

    # Corde is a gendered duo relay (RELAYCOUNT=2), not an individual event.
    if label == "Corde":
        uid = TICKET_UID.get(("Corde", True, False))
        if uid is None:
            return None
        return EventKey(age_code, gender, uid, is_relay=True)

    lookup_label = label
    is_masters_obstacle = False
    # Portage has two variants: 50 m and 100 m.
    if label == "Portage" and dist_txt == "50":
        lookup_label = "Portage50"
    # Obstacle 100 m is only valid for Masters (Masters uses UID 541).
    elif label == "Obstacle" and dist_txt == "100":
        if age_code != "MASTERS":
            # Someone tagged a 15-18 or Open entry with an Obstacle 100m
            # ticket; ignore the distance, map to UID 501 (200 m) but we
            # still report it as a bad combination via validation.
            pass
        else:
            is_masters_obstacle = True

    # Masters Obstacle defaults to 100 m even without an explicit "100 m"
    # suffix, since the template only has UID 541 for Masters.  But be
    # strict: only honour the 100 m variant when both age=MASTERS AND the
    # ticket says "100".  A plain "MA F Obstacle" without distance will
    # still resolve to UID 501 (which the template doesn't have Masters
    # brackets for) and therefore fail validation — the organiser can
    # fix the xlsx.
    if label == "Obstacle" and age_code == "MASTERS" and is_masters_obstacle:
        uid = TICKET_UID.get(("Obstacle", False, True))
    else:
        uid = TICKET_UID.get((lookup_label, False, False))
    if uid is None:
        return None
    return EventKey(age_code, gender, uid, is_relay=False)


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def norm_key(*parts: Any) -> str:
    """Normalise a string to act as a dedup key (lower, strip accents)."""
    s = " ".join((str(p) if p is not None else "") for p in parts).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)


def truncate(s: str | None, n: int) -> str | None:
    if s is None:
        return None
    s = str(s)
    return s[:n]


def upper_key(s: str | None, n: int) -> str | None:
    if s is None:
        return None
    return unicodedata.normalize("NFKD", s).upper().encode("ascii", "ignore") \
        .decode("ascii")[:n]


def age_at(birthdate: dt.datetime | dt.date | None,
           ref: dt.date | None = None) -> int | None:
    """Age in whole years at `ref`, or None if no birthdate."""
    if birthdate is None:
        return None
    if ref is None:
        ref = AGE_DATE
    if ref is None:
        return None
    bd = birthdate.date() if isinstance(birthdate, dt.datetime) else birthdate
    years = ref.year - bd.year
    if (ref.month, ref.day) < (bd.month, bd.day):
        years -= 1
    return years


def pick_agegroup(agegroups: list[tuple],
                  athlete_age: int | None,
                  fallback_min: int, fallback_max: int) -> int | None:
    """From a list of (AGEGROUPID, AGEMIN, AGEMAX), pick the one that best
    contains the athlete's age.  Retained for reference / potential reuse
    (not called directly in the current per-age-bracket model, which uses
    event_targets routing instead).
    """
    if not agegroups:
        return None

    def in_bounds(amin, amax, x):
        lo = -10**9 if amin is None or amin < 0 else amin
        hi = 10**9  if amax is None or amax < 0 else amax
        return lo <= x <= hi

    if athlete_age is not None:
        matches = [(agid, amin, amax) for agid, amin, amax in agegroups
                   if in_bounds(amin, amax, athlete_age)]
        if matches:
            def span(t):
                _, amin, amax = t
                lo = 0   if amin is None or amin < 0 else amin
                hi = 999 if amax is None or amax < 0 else amax
                return hi - lo
            matches.sort(key=span)
            return matches[0][0]
    for agid, amin, amax in agegroups:
        if amin == fallback_min and amax == fallback_max:
            return agid
    for agid, amin, amax in agegroups:
        lo = -10**9 if amin is None or amin < 0 else amin
        hi = 10**9  if amax is None or amax < 0 else amax
        if lo <= fallback_min and fallback_max <= hi:
            return agid
    return agegroups[0][0]


def short_code_from_name(name: str, length: int = 10) -> str:
    """Build a club short code: keep capitals+digits, fall back to prefix."""
    caps = "".join(c for c in name if c.isupper() or c.isdigit())
    if 2 <= len(caps) <= length:
        return caps
    cleaned = re.sub(r"[^A-Za-z0-9]", "", unicodedata.normalize("NFKD", name)
                     .encode("ascii", "ignore").decode("ascii"))
    return cleaned[:length].upper() or "CLUB"


# --------------------------------------------------------------------------- #
# Fuzzy duplicate detection helpers
# --------------------------------------------------------------------------- #
import difflib as _difflib

FUZZY_CLUB_THRESHOLD    = 0.90   # club name similarity
FUZZY_ATHLETE_THRESHOLD = 0.90   # same-club athlete full-name similarity


def fuzzy_key(s: str) -> str:
    """Strong normalisation for dedup: lowercase, NFKD, strip accents and
    punctuation, collapse whitespace.  Two strings with the same
    fuzzy_key are almost certainly the same entity."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s)     # drop punctuation
    s = re.sub(r"\s+", " ", s).strip()
    return s


def similarity(a: str, b: str) -> float:
    """Ratio in [0, 1].  1.0 = identical, 0.0 = unrelated."""
    return _difflib.SequenceMatcher(None, a, b).ratio()


def find_fuzzy_club_duplicates(
        club_counts: dict[str, int],
        threshold: float = FUZZY_CLUB_THRESHOLD) -> list[tuple[str, str, float, int, int]]:
    """Return list of (name_a, name_b, similarity, count_a, count_b) for
    club-name pairs that look like typos/variants of each other.

    Pairs are reported when either:
      - their fuzzy_key is identical (e.g. 'Rouville SurfClub' / 'Rouville Surfclub'), or
      - their similarity on normalised text is >= threshold.
    """
    names = sorted(club_counts.keys(), key=str.lower)
    out: list[tuple[str, str, float, int, int]] = []
    for i in range(len(names)):
        a = names[i]
        ka = fuzzy_key(a)
        for j in range(i + 1, len(names)):
            b = names[j]
            kb = fuzzy_key(b)
            if not ka or not kb:
                continue
            if ka == kb:
                out.append((a, b, 1.0, club_counts[a], club_counts[b]))
                continue
            # Skip comparisons with wildly different lengths — saves time and
            # avoids false positives on very short names.
            if abs(len(ka) - len(kb)) > max(4, min(len(ka), len(kb)) // 2):
                continue
            s = similarity(ka, kb)
            if s >= threshold:
                out.append((a, b, s, club_counts[a], club_counts[b]))
    return out


def find_fuzzy_athlete_duplicates(
        athletes: dict[tuple, "Inscription"],
        threshold: float = FUZZY_ATHLETE_THRESHOLD
) -> dict[str, list[tuple]]:
    """Scan the athletes dict for suspect duplicates.

    Returns a dict with three keys:
        'same_license':          pairs sharing a LICENSE but with different names
        'same_club_fuzzy':       pairs in the same club whose full name is similar
        'cross_club_same_person':pairs in different clubs with same normalised
                                 first+last name AND same birthdate
    """
    results = {
        "same_license":           [],
        "same_club_fuzzy":        [],
        "cross_club_same_person": [],
    }

    # Index athletes by license
    by_license: dict[str, list[tuple[tuple, "Inscription"]]] = {}
    for akey, ins in athletes.items():
        lic = (ins.license or "").strip()
        if lic:
            by_license.setdefault(lic, []).append((akey, ins))

    # 1) same license, different name
    for lic, group in by_license.items():
        if len(group) < 2:
            continue
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                a_key, a = group[i]
                b_key, b = group[j]
                name_a = f"{a.first} {a.last}"
                name_b = f"{b.first} {b.last}"
                if fuzzy_key(name_a) != fuzzy_key(name_b):
                    results["same_license"].append(
                        (name_a, a.club, name_b, b.club, lic))

    # Group athletes by club for same-club fuzzy check
    by_club: dict[str, list[tuple[tuple, "Inscription"]]] = {}
    for akey, ins in athletes.items():
        by_club.setdefault(norm_key(ins.club), []).append((akey, ins))

    # 2) same club, similar full names
    for cnorm, group in by_club.items():
        n = len(group)
        if n < 2:
            continue
        keys = [fuzzy_key(f"{ins.first} {ins.last}") for _, ins in group]
        for i in range(n):
            ka = keys[i]
            if not ka:
                continue
            for j in range(i + 1, n):
                kb = keys[j]
                if not kb or ka == kb:
                    # Exact match after normalisation — same person twice.
                    # (Usually caught upstream by the dedup key, but guard
                    # for cases where license differs so the dedup key
                    # doesn't coalesce them.)
                    if ka == kb:
                        a_key, a = group[i]; b_key, b = group[j]
                        if a.license != b.license:
                            results["same_club_fuzzy"].append(
                                (f"{a.first} {a.last}", a.license or "-",
                                 f"{b.first} {b.last}", b.license or "-",
                                 a.club, 1.0))
                    continue
                if abs(len(ka) - len(kb)) > max(4, min(len(ka), len(kb)) // 2):
                    continue
                s = similarity(ka, kb)
                if s >= threshold:
                    # Extra guard: require BOTH first AND last to be reasonably
                    # similar (so "Alice Tremblay" vs "Alice Gauthier" doesn't
                    # trip the trigger just on the shared first name).
                    a_key, a = group[i]; b_key, b = group[j]
                    s_first = similarity(fuzzy_key(a.first), fuzzy_key(b.first))
                    s_last  = similarity(fuzzy_key(a.last),  fuzzy_key(b.last))
                    if s_first >= 0.70 and s_last >= 0.70:
                        results["same_club_fuzzy"].append(
                            (f"{a.first} {a.last}", a.license or "-",
                             f"{b.first} {b.last}", b.license or "-",
                             a.club, s))

    # 3) cross-club: same normalised name + same birthdate
    by_name_dob: dict[tuple, list[tuple[tuple, "Inscription"]]] = {}
    for akey, ins in athletes.items():
        bd = ins.birthdate
        bd_key = bd.date() if isinstance(bd, dt.datetime) else bd
        k = (fuzzy_key(f"{ins.first} {ins.last}"), bd_key)
        if k[0] and k[1] is not None:
            by_name_dob.setdefault(k, []).append((akey, ins))
    for (name_k, bd_k), group in by_name_dob.items():
        # Skip when everyone is in the same club (already handled above)
        clubs_ = {norm_key(ins.club) for _, ins in group}
        if len(clubs_) < 2:
            continue
        # Report each cross-club pair once
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                _, a = group[i]; _, b = group[j]
                if norm_key(a.club) == norm_key(b.club):
                    continue
                results["cross_club_same_person"].append(
                    (f"{a.first} {a.last}", a.club, b.club,
                     bd_k.isoformat() if bd_k else ""))

    return results


_TIME_RE_H = re.compile(r"^\s*(\d+):(\d{1,2}):(\d{1,2})(?:[.,](\d{1,3}))?\s*$")
_TIME_RE_M = re.compile(r"^\s*(\d+):(\d{1,2})(?:[.,](\d{1,3}))?\s*$")
_TIME_RE_S = re.compile(r"^\s*(\d+)(?:[.,](\d{1,3}))?\s*$")


def parse_best_time(val: Any) -> int | None:
    """Parse a swim best time into MILLISECONDS, or None.

    SPLASH's SWIMRESULT.ENTRYTIME / SWIMTIME column stores milliseconds,
    not hundredths of a second — confirmed against the 30-Deux federation
    sample where a 50 m obstacle result of SWIMTIME=45520 corresponds
    to 45.520 s (as milliseconds), not 7:35.20 (as centiseconds).

    Recognised formats:
        hh:mm:ss[.fff]
        mm:ss[.fff]          <-- most common in lifesaving meets
        ss[.fff]
        An Excel time object (datetime.time / timedelta-like)
        A float/int in seconds
    Fractional part is padded to 3 digits ('1:05' -> 65000 ms,
    '1:05.3' -> 65300 ms, '1:05.37' -> 65370 ms, '1:05.371' -> 65371 ms).
    """
    if val is None:
        return None
    if isinstance(val, dt.time):
        return ((val.hour * 3600 + val.minute * 60 + val.second) * 1000
                + val.microsecond // 1000)
    if isinstance(val, dt.timedelta):
        return int(round(val.total_seconds() * 1000))
    if isinstance(val, (int, float)):
        x = float(val)
        if 0 < x < 1:                    # Excel fraction-of-day
            return int(round(x * 24 * 3600 * 1000))
        if x > 0:                        # seconds
            return int(round(x * 1000))
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nt", "n/a", "na", "-"):
        return None

    def _ms(fs: str | None) -> int:
        """Normalise a fractional-seconds string to milliseconds (0-999)."""
        fs = (fs or "0")
        fs = (fs + "000")[:3]
        return int(fs)

    m = _TIME_RE_H.match(s)
    if m:
        h, mm, ss, fs = m.group(1), m.group(2), m.group(3), m.group(4)
        total = (int(h) * 3600 + int(mm) * 60 + int(ss)) * 1000 + _ms(fs)
        return total if total > 0 else None
    m = _TIME_RE_M.match(s)
    if m:
        mm, ss, fs = m.group(1), m.group(2), m.group(3)
        total = (int(mm) * 60 + int(ss)) * 1000 + _ms(fs)
        return total if total > 0 else None
    m = _TIME_RE_S.match(s)
    if m:
        ss, fs = m.group(1), m.group(2)
        total = int(ss) * 1000 + _ms(fs)
        return total if total > 0 else None
    return None


def parse_birthdate(val: Any) -> dt.datetime | None:
    if val is None or (isinstance(val, float) and val != val):
        return None
    if isinstance(val, dt.datetime):
        return val
    if isinstance(val, dt.date):
        return dt.datetime(val.year, val.month, val.day)
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y",
                "%B %d, %Y", "%b %d, %Y", "%d %B %Y"):
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# --------------------------------------------------------------------------- #
# Excel reader
# --------------------------------------------------------------------------- #
@dataclass
class Inscription:
    first: str
    last: str
    email: str | None
    club: str
    birthdate: dt.datetime | None
    license: str | None
    best_time_ms: int | None
    event: EventKey
    teammates: str | None = None  # raw string, used only for relay debug


# --------------------------------------------------------------------------- #
# Issue collector — data-quality warnings surfaced at end of run
# --------------------------------------------------------------------------- #
@dataclass
class Issue:
    severity: str     # 'WARNING' | 'NOTE'
    category: str     # short tag (e.g. 'bad_time', 'no_dob', 'age_mismatch')
    message: str      # human-readable description
    row: int | None = None      # xlsx row number (1-based) if applicable


class IssueCollector:
    """Bucket for data-quality findings.  At end of run we print a summary
    grouped by category, capped at N items per category."""

    def __init__(self, max_per_category: int = 10):
        self.issues: list[Issue] = []
        self.max_per_category = max_per_category

    def add(self, severity: str, category: str, message: str,
            row: int | None = None) -> None:
        self.issues.append(Issue(severity, category, message, row))

    def warn(self, category: str, message: str, row: int | None = None):
        self.add("WARNING", category, message, row)

    def note(self, category: str, message: str, row: int | None = None):
        self.add("NOTE", category, message, row)

    def by_category(self):
        from collections import defaultdict
        out: dict[tuple, list[Issue]] = defaultdict(list)
        for i in self.issues:
            out[(i.severity, i.category)].append(i)
        return out

    def report(self, title: str = "Issues",
               out_file=None,
               full: bool = False) -> None:
        """Print (and optionally write to `out_file`) the issues section.

        full=True removes the per-category cap (every issue listed)."""
        if not self.issues:
            return
        buckets = self.by_category()
        ordered = sorted(buckets.items(),
                         key=lambda kv: (kv[0][0] != "WARNING", -len(kv[1])))
        cap = 10**9 if full else self.max_per_category

        lines: list[str] = []
        lines.append("")
        lines.append("=" * 60)
        from datetime import datetime, timezone, timedelta
        _et = timezone(timedelta(hours=-4))
        lines.append(f"  {title}  ({datetime.now(_et):%Y-%m-%d %H:%M} ET)")
        lines.append("=" * 60)
        for (sev, cat), items in ordered:
            lines.append(f"  [{sev}] {cat}: {len(items)}")
            for it in items[:cap]:
                suffix = f" (row {it.row})" if it.row else ""
                lines.append(f"       - {it.message}{suffix}")
            if len(items) > cap:
                lines.append(f"       … and {len(items) - cap} more")
        lines.append("=" * 60)

        block = "\n".join(lines)
        print(block)
        if out_file is not None:
            out_file.write(block + "\n")


# --------------------------------------------------------------------------- #
# JotForm matrix format reader
# --------------------------------------------------------------------------- #
# Maps JotForm matrix row labels to TICKET_UID style names
_JOTFORM_STYLE_MAP = {
    "Obstacle": "Obstacle",
    "Remorquage": "Remorquage",
    "Portage (100m)": "Portage",
    "Portage50 (50m)": "Portage50",
    "Sauveteur d'acier": "Sauveteur d'acier",
    "Medley": "Medley",
    "Corde": "Corde",
    "Relais Mixte Obstacle": "Obstacle",
    "Relais Mixte Portage": "Portage",
    "Relais Mixte Medley": "Medley",
}

# Maps JotForm [Col] labels to (age_code, gender)
_JOTFORM_CAT_MAP = {
    "15-18 M": ("1518", GENDER_MALE),
    "15-18 F": ("1518", GENDER_FEMALE),
    "Open M": ("OPEN", GENDER_MALE),
    "Open F": ("OPEN", GENDER_FEMALE),
    "Open": ("OPEN", GENDER_MIXED),
    "MA M": ("MASTERS", GENDER_MALE),
    "MA F": ("MASTERS", GENDER_FEMALE),
}

# Relay row labels
_JOTFORM_RELAY_ROWS = {"Relais Mixte Obstacle", "Relais Mixte Portage",
                       "Relais Mixte Medley", "Corde"}


def _read_jotform(wb, ws, header: list[str],
                  issues: IssueCollector | None) -> list[Inscription]:
    """Parse a JotForm matrix-style export into Inscription records."""
    # Find column indices for standard fields
    def _find_col(keywords):
        for i, h in enumerate(header):
            hl = h.lower()
            if all(k in hl for k in keywords):
                return i
        return None

    i_first = _find_col(["first"])
    i_last = _find_col(["last"])
    # JotForm may export full name as single column
    i_fullname = None
    if i_first is None or i_last is None:
        i_fullname = _find_col(["nom"]) or _find_col(["name"])
        if i_fullname is None:
            i_fullname = _find_col(["athlète"]) or _find_col(["athlete"])
    i_email = _find_col(["courriel"]) or _find_col(["email"])
    i_club = _find_col(["club"])
    i_dob = _find_col(["naissance"]) or _find_col(["birth"])
    i_nran = _find_col(["nran"])
    i_team = _find_col(["coéquipier"]) or _find_col(["teammate"])

    # Parse matrix columns: header contains "[Row][Col]" or ">> Row >> Col"
    # Build mapping: col_index -> (style_name, age_code, gender, is_relay)
    matrix_cols: dict[int, tuple] = {}
    _re_matrix = re.compile(r"\[([^\]]+)\]\[([^\]]+)\]$")
    _re_matrix_single = re.compile(r"\[([^\]]+)\]$")
    _re_matrix_gg = re.compile(r">>\s*([^>]+?)\s*>>\s*([^>]+?)\s*$")
    for i, h in enumerate(header):
        row_label = col_label = None
        m = _re_matrix.search(h)
        if m:
            row_label, col_label = m.group(1).strip(), m.group(2).strip()
        else:
            m = _re_matrix_gg.search(h)
            if m:
                row_label, col_label = m.group(1).strip(), m.group(2).strip()
            else:
                m2 = _re_matrix_single.search(h)
                if m2:
                    row_label = m2.group(1).strip()
                    col_label = "Open"
        if row_label is None:
            continue

        style = _JOTFORM_STYLE_MAP.get(row_label)
        cat = _JOTFORM_CAT_MAP.get(col_label)
        if style is None or cat is None:
            continue
        age_code, gender = cat
        is_relay = row_label in _JOTFORM_RELAY_ROWS
        # For Corde, it's a relay (duo)
        if row_label == "Corde":
            is_relay = True

        # Look up the TICKET_UID
        if is_relay and row_label != "Corde":
            uid = TICKET_UID.get((style, True, False))
        elif row_label == "Corde":
            uid = TICKET_UID.get(("Corde", True, False))
        elif age_code == "MASTERS" and style == "Obstacle":
            uid = TICKET_UID.get(("Obstacle", False, True))
        else:
            uid = TICKET_UID.get((style, False, False))

        if uid is None:
            continue
        matrix_cols[i] = (style, age_code, gender, is_relay, uid)

    out: list[Inscription] = []
    for row_idx, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not r:
            continue
        # Parse name
        if i_fullname is not None:
            fullname = str(r[i_fullname] or "").strip() if i_fullname < len(r) else ""
            if not fullname:
                continue
            parts = fullname.rsplit(" ", 1)
            first = parts[0] if len(parts) > 1 else fullname
            last = parts[1] if len(parts) > 1 else ""
        else:
            first = str(r[i_first] or "").strip() if i_first is not None and i_first < len(r) else ""
            last = str(r[i_last] or "").strip() if i_last is not None and i_last < len(r) else ""
        if not first and not last:
            continue
        email = str(r[i_email]).strip() if i_email is not None and r[i_email] else None
        club = str(r[i_club]).strip() if i_club is not None and r[i_club] else "Unattached"
        raw_dob = r[i_dob] if i_dob is not None else None
        bd = parse_birthdate(raw_dob)
        if raw_dob and bd is None and issues:
            issues.warn("bad_birthdate",
                        f"{first} {last}: can't parse DOB {raw_dob!r}",
                        row=row_idx)
        license_val = str(r[i_nran]).strip() if i_nran is not None and r[i_nran] else None
        teammates = str(r[i_team]).strip() if i_team is not None and r[i_team] else None

        # Expand matrix cells into individual Inscription records
        for col_i, (style, age_code, gender, is_relay, uid) in matrix_cols.items():
            cell = r[col_i] if col_i < len(r) else None
            if cell is None or str(cell).strip() == "":
                continue
            raw_time = str(cell).strip()
            best_ms = parse_best_time(raw_time)
            if raw_time.lower() not in ("nt", "n/a", "na", "-") and best_ms is None:
                if issues:
                    issues.warn("bad_time",
                                f"{first} {last} {style}/{age_code}: "
                                f"can't parse time {raw_time!r}",
                                row=row_idx)

            ev = EventKey(age_code=age_code, gender=gender,
                          uniqueid=uid, is_relay=is_relay)
            out.append(Inscription(
                first=first, last=last, email=email, club=club,
                birthdate=bd, license=license_val,
                best_time_ms=best_ms, event=ev,
                teammates=teammates if is_relay else None,
            ))

    return out


def read_attendees(xlsx_path: Path,
                   issues: IssueCollector | None = None) -> list[Inscription]:
    """Parse the Attendees sheet into a list of Inscription records.

    If an IssueCollector is provided, data-quality problems encountered
    while parsing (missing names, unparseable ticket types, bad times,
    bad birthdates, truncated names, duplicate athlete/event pairs) are
    reported into it.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Detect format: JotForm matrix has "[" or ">>" in column headers
    ws = wb.active if "Attendees" not in wb.sheetnames else wb["Attendees"]
    header = [str(c or "") for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if any(("[" in h and "]" in h) or (">>" in h) for h in header):
        return _read_jotform(wb, ws, header, issues)

    # Eventbrite format — requires "Attendees" sheet
    if "Attendees" not in wb.sheetnames:
        raise SystemExit(f"Sheet 'Attendees' not found in {xlsx_path}")
    ws = wb["Attendees"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c or "").strip() for c in rows[0]]
    def col(name): return header.index(name)
    i_first = col("First Name")
    i_last = col("Last Name")
    i_email = col("Email")
    i_ticket = col("Ticket Type")
    i_best = col("Best time")
    i_club = col("Club")
    i_dob = col("DD/MM/YYYY")
    i_lic = col("NRAN")
    try:
        i_team = header.index("Teammate(s) + NRAN")
    except ValueError:
        i_team = None

    out: list[Inscription] = []
    # Track duplicate (athlete, event) pairs from the xlsx itself
    seen_pairs: dict[tuple, tuple[int, int | None]] = {}

    for row_idx, r in enumerate(rows[1:], start=2):
        if not r:
            continue
        first = r[i_first]; last = r[i_last]
        if not first or not last:
            # Skip fully empty rows silently; warn on partial ones.
            if any(cell not in (None, "") for cell in r):
                if issues:
                    issues.warn("missing_name",
                                f"row missing first or last name",
                                row=row_idx)
            continue
        ticket = (str(r[i_ticket] or "").strip()) if r[i_ticket] else ""
        if not ticket:
            if issues:
                issues.warn("missing_ticket",
                            f"{first} {last}: empty Ticket Type",
                            row=row_idx)
            continue
        ev = parse_ticket(ticket)
        if ev is None:
            # Distinguish intentional non-race tickets from unknown ones
            if any(ticket.startswith(p) for p in NON_RACE_PREFIXES):
                # legitimate non-race (Banquet, Coach, etc.)
                pass
            else:
                if issues:
                    issues.warn("unknown_ticket",
                                f"{first} {last}: unrecognised ticket "
                                f"{ticket!r}",
                                row=row_idx)
            continue

        # Parse best time + birthdate; track failures
        raw_time = r[i_best]
        best_ms = parse_best_time(raw_time)
        if raw_time not in (None, "") and best_ms is None \
                and str(raw_time).strip().lower() not in ("nt", "n/a", "na", "-"):
            if issues:
                issues.warn("bad_time",
                            f"{first} {last} {ticket!r}: "
                            f"can't parse time {raw_time!r}",
                            row=row_idx)

        raw_dob = r[i_dob]
        bd = parse_birthdate(raw_dob)
        if raw_dob not in (None, "") and bd is None:
            if issues:
                issues.warn("bad_birthdate",
                            f"{first} {last}: can't parse DOB {raw_dob!r}",
                            row=row_idx)
        elif bd is not None and issues:
            age = age_at(bd)
            if age is not None and (age < 0 or age > 99):
                issues.warn("bad_birthdate",
                            f"{first} {last}: implausible age {age} "
                            f"(DOB {bd.strftime('%Y-%m-%d')})",
                            row=row_idx)

        # Name-length warnings (will be truncated for SPLASH columns)
        if len(str(first)) > 30 and issues:
            issues.note("truncated_name",
                        f"first name truncated (>30 chars): {first!r}",
                        row=row_idx)
        if len(str(last)) > 50 and issues:
            issues.note("truncated_name",
                        f"last name truncated (>50 chars): {last!r}",
                        row=row_idx)
        club_raw = (r[i_club] or "Unattached")
        if len(str(club_raw)) > 80 and issues:
            issues.note("truncated_name",
                        f"club name truncated (>80 chars): {club_raw!r}",
                        row=row_idx)

        # Duplicate (athlete, event) in xlsx — we'll keep the best time, but
        # let the user know so they can clean up the source sheet.
        pair_key = (norm_key(first, last), r[i_lic] or "", ev.key())
        if pair_key in seen_pairs and issues:
            prev_row, prev_cs = seen_pairs[pair_key]
            issues.note("duplicate_entry",
                        f"{first} {last} entered in {ticket!r} again "
                        f"(first seen row {prev_row}); keeping best time",
                        row=row_idx)
        seen_pairs[pair_key] = (row_idx, best_ms)

        out.append(Inscription(
            first=str(first).strip(),
            last=str(last).strip(),
            email=(r[i_email] or None),
            club=str(club_raw).strip(),
            birthdate=bd,
            license=(str(r[i_lic]).strip() if r[i_lic] else None),
            best_time_ms=best_ms,
            event=ev,
            teammates=(str(r[i_team]).strip()
                       if i_team is not None and r[i_team] else None),
        ))
    return out


# --------------------------------------------------------------------------- #
# MDB writer
# --------------------------------------------------------------------------- #
class MDB:
    """Thin wrapper around a UCanAccess JDBC connection."""

    def __init__(self, path: Path, dry_run: bool = False):
        self.path = path
        self.dry_run = dry_run
        jars = (
            glob.glob(os.path.join(UCANACCESS_DIR, "ucanaccess-*.jar"))
            + glob.glob(os.path.join(UCANACCESS_DIR, "lib", "*.jar"))
            # Support the flattened layout used by the Docker image,
            # where all five jars live directly under /opt/ucanaccess/
            # with no lib/ subdir.
            + glob.glob(os.path.join(UCANACCESS_DIR, "*.jar"))
        )
        # Dedup while preserving order
        seen = set()
        jars = [j for j in jars if not (j in seen or seen.add(j))]
        if not jars:
            raise SystemExit(
                f"UCanAccess jars not found under {UCANACCESS_DIR}. "
                "Set UCANACCESS_DIR env var."
            )
        url = (f"jdbc:ucanaccess://{path};openExclusive=false;"
               "memory=true;ignoreCase=true")
        self.conn = jaydebeapi.connect(
            "net.ucanaccess.jdbc.UcanaccessDriver", url, [], jars)
        self.cur = self.conn.cursor()
        self._uid = self._read_uid()
        self._start_uid = self._uid
        # JPype is started by jaydebeapi.connect(); grab Timestamp class
        import jpype  # noqa
        self._Timestamp = jpype.JClass("java.sql.Timestamp")

    # ----- UID allocator -----
    def _read_uid(self) -> int:
        self.cur.execute("SELECT LASTUID FROM BSUIDTABLE WHERE NAME='BS_GLOBAL_UID'")
        return int(self.cur.fetchone()[0])

    def next_id(self) -> int:
        self._uid += 1
        return self._uid

    def flush_uid(self):
        self.cur.execute(
            "UPDATE BSUIDTABLE SET LASTUID=? WHERE NAME='BS_GLOBAL_UID'",
            [self._uid])

    # ----- DML -----
    def exec(self, sql: str, params: list | None = None):
        if self.dry_run:
            return
        self.cur.execute(sql, params or [])

    def exec_many(self, sql: str, batch: list[list]):
        if self.dry_run or not batch:
            return
        self.cur.executemany(sql, batch)

    def insert(self, table: str, row: dict):
        """INSERT with only non-None columns – avoids UCanAccess NPE on nulls."""
        if self.dry_run:
            return
        cols = [c for c, v in row.items() if v is not None]
        vals = [self._to_jdbc(row[c]) for c in cols]
        if not cols:
            return
        placeholders = ",".join("?" * len(cols))
        sql = (f"INSERT INTO {table} ({','.join(cols)}) "
               f"VALUES ({placeholders})")
        self.cur.execute(sql, vals)

    def insert_many(self, table: str, rows: list[dict]):
        if self.dry_run or not rows:
            return
        # group by the set of non-null columns so each batch has homogeneous SQL
        from itertools import groupby
        def key(r): return tuple(c for c, v in r.items() if v is not None)
        rows_sorted = sorted(rows, key=key)
        for cols, grp in groupby(rows_sorted, key=key):
            cols = list(cols)
            if not cols:
                continue
            placeholders = ",".join("?" * len(cols))
            sql = (f"INSERT INTO {table} ({','.join(cols)}) "
                   f"VALUES ({placeholders})")
            batch = [[self._to_jdbc(r[c]) for c in cols] for r in grp]
            self.cur.executemany(sql, batch)

    def update(self, table: str, where: dict, updates: dict):
        """UPDATE with non-None updates only, filtered by `where` dict."""
        if self.dry_run or not updates:
            return
        set_cols = list(updates.keys())
        where_cols = list(where.keys())
        sql = (f"UPDATE {table} SET "
               + ", ".join(f"{c}=?" for c in set_cols)
               + " WHERE "
               + " AND ".join(f"{c}=?" for c in where_cols))
        params = ([self._to_jdbc(updates[c]) for c in set_cols]
                  + [self._to_jdbc(where[c])  for c in where_cols])
        self.cur.execute(sql, params)

    def query(self, sql: str, params: list | None = None):
        """SELECT and return all rows as list[tuple]."""
        self.cur.execute(sql, params or [])
        return self.cur.fetchall()

    def _to_jdbc(self, v):
        """Convert Python value into something JDBC's setObject accepts."""
        if isinstance(v, dt.datetime):
            # java.sql.Timestamp(long ms since epoch)
            epoch_ms = int(v.timestamp() * 1000)
            return self._Timestamp(epoch_ms)
        if isinstance(v, dt.date):
            epoch_ms = int(dt.datetime(v.year, v.month, v.day).timestamp() * 1000)
            return self._Timestamp(epoch_ms)
        return v

    def commit(self):
        if self.dry_run:
            print("[dry-run] skipping commit; rolling back")
            self.conn.rollback()
        else:
            self.flush_uid()
            self.conn.commit()

    def close(self):
        self.conn.close()


# --------------------------------------------------------------------------- #
# Main loader
# --------------------------------------------------------------------------- #
# --------------------------------------------------------------------------- #
# TemplateIndex — authoritative event structure from the supplied MDB
# --------------------------------------------------------------------------- #
@dataclass
class TemplateStyle:
    swim_style_id: int
    uniqueid: int
    name: str | None
    distance: int | None
    relay_count: int

@dataclass
class TemplateAgeGroup:
    agegroup_id: int
    amin: int | None
    amax: int | None
    gender: int | None

@dataclass
class TemplateEvent:
    swim_event_id: int
    swim_style_id: int
    uniqueid: int
    gender: int
    round: int
    event_number: int | None
    session_id: int | None
    masters: bool
    agegroups: list[TemplateAgeGroup]


class TemplateIndex:
    """Snapshot of the supplied .mdb's event structure.

    Built once before any write.  Provides lookups:
      - styles_by_uid  : UNIQUEID -> TemplateStyle
      - events_by_uid_gender : (UNIQUEID, GENDER) -> list[TemplateEvent]
        (typically 2 entries: prelim + final; Masters events are single
         timed-finals)
      - is_first_run   : True iff zero SWIMRESULT + zero RELAY rows
    """

    def __init__(self, db: "MDB"):
        # Styles — only interested in lifesaving-catalog rows (STROKE=0).
        self.styles_by_uid: dict[int, TemplateStyle] = {}
        for sid, uid, name, dist, rc, stroke in db.query(
            "SELECT SWIMSTYLEID, UNIQUEID, NAME, DISTANCE, RELAYCOUNT, STROKE "
            "FROM SWIMSTYLE"):
            if uid is None:
                continue
            uid_i = int(uid)
            # We don't filter by stroke=0 here — the template has some
            # swim strokes too, but our ticket UIDs all fall in the 500+
            # range and thus never collide.
            self.styles_by_uid[uid_i] = TemplateStyle(
                swim_style_id=int(sid), uniqueid=uid_i,
                name=name,
                distance=int(dist) if dist is not None else None,
                relay_count=int(rc) if rc is not None else 1)

        # Age groups grouped by SWIMEVENTID
        ag_by_event: dict[int, list[TemplateAgeGroup]] = defaultdict(list)
        for agid, seid, amin, amax, gen in db.query(
            "SELECT AGEGROUPID, SWIMEVENTID, AGEMIN, AGEMAX, GENDER "
            "FROM AGEGROUP"):
            if seid is None:
                continue
            ag_by_event[int(seid)].append(TemplateAgeGroup(
                agegroup_id=int(agid),
                amin=int(amin) if amin is not None else None,
                amax=int(amax) if amax is not None else None,
                gender=int(gen) if gen is not None else None))

        # Events
        self.events_by_uid_gender: dict[tuple, list[TemplateEvent]] = defaultdict(list)
        for eid, styid, gen, rnd, enum, ses, mas in db.query(
            "SELECT SWIMEVENTID, SWIMSTYLEID, GENDER, ROUND, EVENTNUMBER, "
            "       SWIMSESSIONID, MASTERS FROM SWIMEVENT"):
            if styid is None or gen is None:
                continue
            styid_i = int(styid)
            # Find the UNIQUEID for this SWIMSTYLEID
            uid = None
            for s in self.styles_by_uid.values():
                if s.swim_style_id == styid_i:
                    uid = s.uniqueid; break
            if uid is None:
                continue
            ev = TemplateEvent(
                swim_event_id=int(eid), swim_style_id=styid_i,
                uniqueid=uid, gender=int(gen),
                round=int(rnd) if rnd is not None else 0,
                event_number=int(enum) if enum is not None else None,
                session_id=int(ses) if ses is not None else None,
                masters=(mas == "T"),
                agegroups=ag_by_event.get(int(eid), []))
            self.events_by_uid_gender[(uid, int(gen))].append(ev)

        # Pre-existing inscriptions → drives "first run vs re-run" detection
        self._has_results = any(True for _ in db.query(
            "SELECT TOP 1 SWIMRESULTID FROM SWIMRESULT"))
        if not self._has_results:
            self._has_results = any(True for _ in db.query(
                "SELECT TOP 1 RELAYID FROM RELAY"))

        # Read AGEDATE from MEETVALUES in BSGLOBAL
        global AGE_DATE
        for (data,) in db.query(
                "SELECT DATA FROM BSGLOBAL WHERE NAME='MEETVALUES'"):
            m = re.search(r"AGEDATE=D;(\d{8})", data or "")
            if m:
                AGE_DATE = dt.date(int(m.group(1)[:4]),
                                   int(m.group(1)[4:6]),
                                   int(m.group(1)[6:8]))
                break
        # Also set on the module by name (handles __main__ vs import)
        import sys
        mod = sys.modules.get("load_to_mdb")
        if mod is not None and mod is not sys.modules.get(__name__):
            mod.AGE_DATE = AGE_DATE

    @property
    def is_first_run(self) -> bool:
        return not self._has_results

    def find_event(self, uid: int, gender: int, masters: bool
                    ) -> TemplateEvent | None:
        """Pick the SWIMEVENT for (uid, gender) most appropriate for the
        ticket's age bracket.  For 15-18 / Open we want ROUND=2 (prelim);
        for Masters we want ROUND=1 (timed final).  If the preferred
        round isn't available, fall back to any event for this (uid, gen).
        """
        candidates = self.events_by_uid_gender.get((uid, gender), [])
        if not candidates:
            return None
        if masters:
            ms = [e for e in candidates if e.masters or e.round == ROUND_TIMED_FINAL]
            if ms:
                return ms[0]
        else:
            # Prefer prelim, then timed-final, then first
            for r in (ROUND_PRELIM, ROUND_TIMED_FINAL, ROUND_FINAL):
                for e in candidates:
                    if e.round == r and not e.masters:
                        return e
        return candidates[0]

    def find_prelim_for_dual_entry(self, uid: int, gender: int
                                    ) -> TemplateEvent | None:
        """Find the non-Masters prelim event for (uid, gender) that has
        at least one Masters-style age bracket (amin in 25..99).
        Returns None if no such event exists (dual-entry not configured).
        """
        candidates = self.events_by_uid_gender.get((uid, gender), [])
        for e in candidates:
            if e.masters or e.round != ROUND_PRELIM:
                continue
            for a in e.agegroups:
                if a.amin is not None and 25 <= a.amin < 100:
                    return e
        return None


def pick_agegroup_for_individual(
        event: TemplateEvent, age_code: str, athlete_age: int | None
) -> TemplateAgeGroup | None:
    """Pick the AGEGROUP within `event` that matches the ticket's age
    bracket.  Returns None if no match."""
    if age_code == "1518":
        for a in event.agegroups:
            if a.amin == 15 and a.amax == 18:
                return a
    elif age_code == "OPEN":
        # '19 & over' — AGEMAX is 99, -1 or None
        for a in event.agegroups:
            if a.amin == 19 and (a.amax in (99, -1, None)):
                return a
    elif age_code == "MASTERS":
        if athlete_age is None:
            return None
        for a in event.agegroups:
            if a.amin is None or a.amax is None:
                continue
            # skip the 15-18 and 19-99 brackets if present on the same event
            if (a.amin, a.amax) in ((15, 18), (19, 99), (19, -1)):
                continue
            lo = a.amin
            hi = 10**9 if a.amax < 0 else a.amax
            if lo <= athlete_age <= hi:
                return a
    return None


def pick_agegroup_for_relay(
        event: TemplateEvent, age_code: str, squad_age_sum: int | None,
        oldest_age: int | None = None
) -> TemplateAgeGroup | None:
    """Pick the AGEGROUP for a relay in `event`.  15-18/Open use their
    named bracket; Masters relays route by total-age-sum to the matching
    age-sum bracket (amin >= 100), OR by the oldest member's individual
    age if the event only has individual-style brackets (amin < 100,
    e.g. Corde duo)."""
    if age_code == "1518":
        for a in event.agegroups:
            if a.amin == 15 and a.amax == 18:
                return a
    elif age_code == "OPEN":
        for a in event.agegroups:
            if a.amin == 19 and (a.amax in (99, -1, None)):
                return a
    elif age_code == "MASTERS":
        # Determine if this event uses age-sum brackets (amin >= 100)
        # or individual-style brackets (amin in 25..99).
        has_agesum = any(a.amin is not None and a.amin >= 100
                         for a in event.agegroups)
        if has_agesum:
            if squad_age_sum is None:
                return None
            for a in event.agegroups:
                if a.amin is None or a.amax is None:
                    continue
                if a.amin in (15, 19):
                    continue
                lo = a.amin
                hi = 10**9 if a.amax < 0 else a.amax
                if lo <= squad_age_sum <= hi:
                    return a
        else:
            # Individual-style brackets — route by oldest member's age
            if oldest_age is None:
                return None
            for a in event.agegroups:
                if a.amin is None or a.amax is None:
                    continue
                if a.amin in (15, 19):
                    continue
                lo = a.amin
                hi = 10**9 if a.amax < 0 else a.amax
                if lo <= oldest_age <= hi:
                    return a
    return None


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--mdb", required=True, type=Path,
                    help="Target .mdb file (will be modified).")
    ap.add_argument("--dry-run", action="store_true",
                    help="Parse and plan everything but rollback at the end.")
    ap.add_argument("--issues-full", action="store_true",
                    help="List every issue, no per-category cap. "
                         "Useful when handing the report to someone who "
                         "has to fix the source xlsx row by row.")
    ap.add_argument("--issues-out", type=Path,
                    help="Also write the issues section to the given "
                         "file (plain text). Implies --issues-full.")
    args = ap.parse_args()

    if args.issues_out:
        args.issues_full = True

    if not args.xlsx.exists():
        sys.exit(f"xlsx not found: {args.xlsx}")
    if not args.mdb.exists():
        sys.exit(f"mdb not found: {args.mdb}")

    print(f"Reading {args.xlsx}...")
    issues = IssueCollector()
    inscriptions = read_attendees(args.xlsx, issues)
    print(f"  {len(inscriptions)} race inscriptions extracted")

    # Aggregate
    from common import aggregate, run_sanity_checks, run_validation, run_cross_row_checks
    data = aggregate(inscriptions, issues)
    clubs = data.clubs
    athletes = data.athletes
    name_to_key = data.name_to_key
    events_in_xlsx = data.events_in_xlsx
    ind_entries = data.ind_entries
    relay_squads = data.relay_squads

    print(f"  distinct clubs:    {len(clubs)}")
    print(f"  distinct athletes: {len(athletes)}")
    print(f"  distinct tickets:  {len(events_in_xlsx)}  "
          f"(individual: {sum(1 for e in events_in_xlsx.values() if not e.is_relay)}, "
          f"relay: {sum(1 for e in events_in_xlsx.values() if e.is_relay)})")
    print(f"  individual entries: {len(ind_entries)}")
    relay_count = sum(len(squads) for squads in relay_squads.values())
    print(f"  relay squads:      {relay_count}")

    # ----- open MDB + load template structure -----
    print(f"\nOpening {args.mdb}...")
    db = MDB(args.mdb, dry_run=args.dry_run)
    print(f"  starting BS_GLOBAL_UID = {db._uid}")

    template = TemplateIndex(db)
    print(f"  age date: {AGE_DATE}")
    n_events_in_template = sum(
        len(v) for v in template.events_by_uid_gender.values())
    print(f"  template: {len(template.styles_by_uid)} SWIMSTYLE rows, "
          f"{n_events_in_template} SWIMEVENTs")

    # Sanity check: all TICKET_UID values must exist in the template
    sanity_errors = run_sanity_checks(template)
    if sanity_errors:
        for e in sanity_errors:
            print(f"\n  FATAL: {e}")
        sys.exit(2)
    if template.is_first_run:
        print(f"  no inscriptions in template — FIRST RUN")
    else:
        print(f"  existing inscriptions detected — UPDATE RE-RUN")

    # ========================================================== #
    # VALIDATION PASS — no writes before this succeeds            #
    # ========================================================== #
    fatal = run_validation(events_in_xlsx, template)

    if fatal:
        print("\n" + "=" * 60)
        print("  FATAL: template/xlsx mismatch — aborting import")
        print("=" * 60)
        for f in fatal:
            print(f"  - {f}")
        print("=" * 60)
        print(f"\n{len(fatal)} fatal error(s).  No writes performed.")
        db.conn.rollback()
        db.close()
        sys.exit(2)

    # ----- Cross-row data-quality checks (warnings only) -----
    run_cross_row_checks(data, template, issues)

    # Helper: infer athlete gender from their individual entries
    athlete_gender_map: dict[tuple, int] = {}
    for akey, ekey, _ in ind_entries:
        ev = events_in_xlsx[ekey]
        if not ev.is_relay and ev.gender in (GENDER_MALE, GENDER_FEMALE):
            athlete_gender_map.setdefault(akey, ev.gender)

    def _infer_gender(akey, ins):
        return athlete_gender_map.get(akey)

    # Gendered relay member gender check
    for (cnorm, ekey), squads in relay_squads.items():
        ev = events_in_xlsx[ekey]
        if ev.gender not in (GENDER_MALE, GENDER_FEMALE):
            continue
        style = template.styles_by_uid[ev.uniqueid]
        need = style.relay_count or 4
        for squad in squads:
            for akey, _ in squad[:need]:
                member = athletes[akey]
                m_gender = _infer_gender(akey, member)
                if m_gender and m_gender != ev.gender:
                    g_label = "M" if ev.gender == GENDER_MALE else "F"
                    issues.warn("relay_member_gender",
                        f"{member.first} {member.last} wrong gender "
                        f"for {g_label} relay ({clubs[cnorm]})")

    # Non-race-only clubs / athletes (informational)
    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if "Attendees" not in wb.sheetnames:
        # JotForm format — skip non-race-only detection
        wb.close()
    else:
        ws_all = wb["Attendees"]
        rows_all = list(ws_all.iter_rows(values_only=True))
        hdr = [str(c or "").strip() for c in rows_all[0]]
        i_f  = hdr.index("First Name")
        i_l  = hdr.index("Last Name")
        i_cl = hdr.index("Club")
        all_clubs: set[str] = set()
        all_names: set[str] = set()
        for r in rows_all[1:]:
            if not r or not r[i_f] or not r[i_l]:
                continue
            all_clubs.add(norm_key(r[i_cl] or "Unattached"))
            all_names.add(norm_key(r[i_f], r[i_l]))
        race_names = {akey[0] for akey in athletes.keys()}
        race_clubs = set(clubs.keys())
        n_club_skipped = len(all_clubs - race_clubs)
        n_ath_skipped = len(all_names - race_names)
        if n_club_skipped:
            issues.note("non_race_only_club",
                        f"{n_club_skipped} club(s) appear only on non-race "
                        f"tickets (Banquet/Coach/Officiel/…) — not imported")
        if n_ath_skipped:
            issues.note("non_race_only_athlete",
                        f"{n_ath_skipped} attendee(s) only bought non-race "
                        f"tickets (supporters, coaches, officials, hotel) "
                        f"— not imported as athletes")
        wb.close()

    # ----- Fuzzy duplicate detection (clubs + athletes) -----
    club_row_counts: dict[str, int] = defaultdict(int)
    for ins in inscriptions:
        club_row_counts[ins.club] += 1
    for a, b, sim, ca, cb in find_fuzzy_club_duplicates(dict(club_row_counts)):
        issues.warn(
            "possible_duplicate_club",
            f"{a!r} ({ca} rows) vs {b!r} ({cb} rows) — similarity {sim:.2f}")
    fuzzy = find_fuzzy_athlete_duplicates(athletes)
    for (name_a, club_a, name_b, club_b, lic) in fuzzy["same_license"]:
        issues.warn(
            "license_name_mismatch",
            f"license {lic!r}: {name_a!r} ({club_a}) vs {name_b!r} ({club_b})"
            f" — same license, different name spelling")
    for (name_a, lic_a, name_b, lic_b, club, sim) in fuzzy["same_club_fuzzy"]:
        issues.warn(
            "possible_duplicate_athlete",
            f"{club}: {name_a!r} (NRAN {lic_a}) vs {name_b!r} "
            f"(NRAN {lic_b}) — similarity {sim:.2f}")
    for (name, club_a, club_b, dob) in fuzzy["cross_club_same_person"]:
        issues.warn(
            "same_person_different_club",
            f"{name!r} born {dob} appears in both {club_a!r} and "
            f"{club_b!r} — probably the same person")

    # ----- Preload existing rows for additive mode -----
    stats = {
        "club_new": 0, "athlete_new": 0,
        "athlete_gender_fix": 0, "athlete_license_fix": 0,
        "athlete_birthdate_fix": 0, "athlete_club_fix": 0,
        "entry_new": 0, "entry_time_faster": 0,
        "relay_new": 0, "relayposition_new": 0,
        "masters_skipped_no_dob": 0,
    }

    existing_clubs: dict[str, tuple[int, str]] = {}
    for cid, name in db.query("SELECT CLUBID, NAME FROM CLUB"):
        if name:
            existing_clubs[norm_key(name)] = (int(cid), name)

    existing_athletes: dict[tuple, dict] = {}
    for aid, clubid, first, last, gender, bdate, lic in db.query(
        "SELECT ATHLETEID, CLUBID, FIRSTNAME, LASTNAME, GENDER, BIRTHDATE, "
        "LICENSE FROM ATHLETE"):
        k = (norm_key(first or "", last or ""), (lic or "").strip())
        existing_athletes[k] = {
            "ATHLETEID": int(aid),
            "CLUBID":    int(clubid) if clubid is not None else None,
            "GENDER":    int(gender) if gender is not None else None,
            "BIRTHDATE": bdate,
            "LICENSE":   lic,
        }

    existing_results: dict[tuple, tuple[int, int | None]] = {}
    for srid, aid, seid, etime in db.query(
        "SELECT SWIMRESULTID, ATHLETEID, SWIMEVENTID, ENTRYTIME FROM SWIMRESULT"):
        if aid is None or seid is None:
            continue
        existing_results[(int(aid), int(seid))] = (
            int(srid), int(etime) if etime is not None else None)

    existing_relay_pos: set[tuple] = set()
    for rid, rnum in db.query(
        "SELECT RELAYID, RELAYNUMBER FROM RELAYPOSITION"):
        if rid is None or rnum is None:
            continue
        existing_relay_pos.add((int(rid), int(rnum)))

    # Per-(club, event, squad-index) stable key for relay dedup
    existing_relays_stable: dict[tuple, int] = {}
    _club_squad_count: dict[tuple, int] = defaultdict(int)
    for rid_row, club_row, event_row, _tn in db.query(
        "SELECT RELAYID, CLUBID, SWIMEVENTID, TEAMNUMBER "
        "FROM RELAY ORDER BY RELAYID"):
        if club_row is None or event_row is None:
            continue
        ce = (int(club_row), int(event_row))
        _club_squad_count[ce] += 1
        existing_relays_stable[(int(club_row), int(event_row),
                                 _club_squad_count[ce])] = int(rid_row)

    rows = db.query("SELECT COALESCE(MAX(TEAMNUMBER), 0) FROM RELAY")
    next_team_no = int(rows[0][0]) if rows and rows[0][0] is not None else 0

    INT_MAX = 2147483647

    # ----- CLUB -----
    club_ids: dict[str, int] = {}
    for cnorm, cname in sorted(clubs.items(), key=lambda kv: kv[1].lower()):
        if cnorm in existing_clubs:
            club_ids[cnorm] = existing_clubs[cnorm][0]
            continue
        cid = db.next_id()
        club_ids[cnorm] = cid
        short = short_code_from_name(cname, 10)
        db.insert("CLUB", {
            "CLUBID":    cid,
            "NAME":      truncate(cname, 80),
            "SHORTNAME": truncate(short, 30),
            "NATION":    MEET_NATION,
            "CLUBTYPE":  1,
            "CODE":      truncate(short, 10),
        })
        stats["club_new"] += 1
        existing_clubs[cnorm] = (cid, cname)

    # ----- ATHLETE -----
    athlete_ids: dict[tuple, int] = {}
    inferred_gender: dict[tuple, int] = {}
    for e in inscriptions:
        if e.event.is_relay:
            continue
        k = (norm_key(e.first, e.last), (e.license or "").strip())
        inferred_gender.setdefault(k, e.event.gender)

    for akey, ins in athletes.items():
        # Skip non-canonical keys (duplicates with different license)
        nk = norm_key(ins.first, ins.last)
        if name_to_key.get(nk) != akey:
            # Map this key to the canonical athlete's ID (set later)
            continue
        club_id = club_ids[norm_key(ins.club)]
        new_gender = inferred_gender.get(akey, GENDER_ALL)
        if akey in existing_athletes:
            existing = existing_athletes[akey]
            aid = existing["ATHLETEID"]
            athlete_ids[akey] = aid
            updates = {}
            if (existing["GENDER"] in (None, GENDER_ALL)
                    and new_gender in (GENDER_MALE, GENDER_FEMALE)):
                updates["GENDER"] = new_gender
                stats["athlete_gender_fix"] += 1
            if not existing["LICENSE"] and ins.license:
                updates["LICENSE"] = truncate(ins.license, 20)
                stats["athlete_license_fix"] += 1
            if existing["BIRTHDATE"] is None and ins.birthdate is not None:
                updates["BIRTHDATE"] = ins.birthdate
                stats["athlete_birthdate_fix"] += 1
            if existing["CLUBID"] != club_id:
                updates["CLUBID"] = club_id
                stats["athlete_club_fix"] += 1
            if updates:
                db.update("ATHLETE", {"ATHLETEID": aid}, updates)
            continue

        aid = db.next_id()
        athlete_ids[akey] = aid
        db.insert("ATHLETE", {
            "ATHLETEID":       aid,
            "CLUBID":          club_id,
            "FIRSTNAME":       truncate(ins.first, 30),
            "LASTNAME":        truncate(ins.last, 50),
            "FIRSTNAME_UPPER": upper_key(ins.first, 5),
            "LASTNAME_UPPER":  upper_key(ins.last, 10),
            "GENDER":          new_gender,
            "BIRTHDATE":       ins.birthdate,
            "LICENSE":         truncate(ins.license, 20),
            "NATION":          MEET_NATION,
            "HANDICAPS":       0,
            "HANDICAPSB":      0,
            "HANDICAPSM":      0,
            "SDMSID":          0,
            "SWRID":           0,
        })
        stats["athlete_new"] += 1
        existing_athletes[akey] = {
            "ATHLETEID": aid, "CLUBID": club_id,
            "GENDER": new_gender, "BIRTHDATE": ins.birthdate,
            "LICENSE": ins.license}

    # Map non-canonical keys to the canonical athlete_id
    for akey in athletes:
        if akey not in athlete_ids:
            nk = norm_key(athletes[akey].first, athletes[akey].last)
            canonical = name_to_key.get(nk)
            if canonical and canonical in athlete_ids:
                athlete_ids[akey] = athlete_ids[canonical]

    # ----- SWIMRESULT (individual entries) -----
    # Dedup (athlete, ek) pairs, keeping fastest time
    best_by: dict[tuple, tuple[int | None, dt.date | None]] = {}
    for akey, ekey, cs in ind_entries:
        ath = athletes[akey]
        cur = best_by.get((akey, ekey))
        if cur is None or (cs is not None and (cur[0] is None or cs < cur[0])):
            best_by[(akey, ekey)] = (cs, ath.birthdate)

    def _sr_row(sr_id, aid, eid, agid, cs):
        return {
            "SWIMRESULTID":  sr_id,
            "ATHLETEID":     aid,
            "SWIMEVENTID":   eid,
            "AGEGROUPID":    agid,
            "ENTRYTIME":     cs,
            "ENTRYCOURSE":   0,
            "RESULTSTATUS":  0,
            "BONUSENTRY":    "F",
            "DSQNOTIFIED":   "F",
            "FINALFIX":      "F",
            "LATEENTRY":     "F",
            "NOADVANCE":     "F",
            "BACKUPTIME1":   None, "BACKUPTIME2": None, "BACKUPTIME3": None,
            "FINISHJUDGE":   None,
            "PADTIME":       None,
            "QTCOURSE":      0,
            "QTTIME":        INT_MAX,
            "QTTIMING":      0,
            "REACTIONTIME":  -32768,
        }

    sr_batch: list[dict] = []
    for (akey, ekey), (cs, bd) in best_by.items():
        ev = events_in_xlsx[ekey]
        athlete_age = age_at(bd)

        # Masters individuals go to the non-Masters prelim (in a Masters
        # bracket) so they swim alongside 15-18/Open.  The Masters final
        # entry is created later by the "Transfert des temps" script.
        if ev.age_code == "MASTERS":
            prelim_ev = template.find_prelim_for_dual_entry(
                ev.uniqueid, ev.gender)
            if prelim_ev is not None:
                tevent = prelim_ev
            else:
                # No prelim with Masters bracket — fall back to Masters final
                tevent = template.find_event(
                    ev.uniqueid, ev.gender, masters=True)
        else:
            tevent = template.find_event(
                ev.uniqueid, ev.gender, masters=False)

        # validation passed, so tevent is guaranteed
        ag = pick_agegroup_for_individual(tevent, ev.age_code, athlete_age)
        if ag is None:
            # Only possible for Masters with no DOB — warn and skip
            if ev.age_code == "MASTERS":
                ins = athletes[akey]
                issues.warn("masters_no_dob",
                    f"{ins.first} {ins.last} Masters entry skipped — "
                    f"no birthdate to route into a 5-year bracket")
                stats["masters_skipped_no_dob"] += 1
            continue

        aid = athlete_ids[akey]
        eid = tevent.swim_event_id

        # Primary entry
        if (aid, eid) in existing_results:
            _sr_id, cur_cs = existing_results[(aid, eid)]
            if cs is not None and (cur_cs is None or cs < cur_cs):
                db.update("SWIMRESULT", {"SWIMRESULTID": _sr_id},
                          {"ENTRYTIME": cs})
                stats["entry_time_faster"] += 1
        else:
            sr_id = db.next_id()
            sr_batch.append(_sr_row(sr_id, aid, eid, ag.agegroup_id, cs))
            stats["entry_new"] += 1
            existing_results[(aid, eid)] = (sr_id, cs)

    db.insert_many("SWIMRESULT", sr_batch)

    # ----- RELAY + RELAYPOSITION -----
    _relay_squad_counter: dict[tuple, int] = {}  # (club_id, event_id) -> next idx
    for (cnorm, ekey), squads in relay_squads.items():
        ev = events_in_xlsx[ekey]
        tevent = template.find_event(
            ev.uniqueid, ev.gender, masters=(ev.age_code == "MASTERS"))
        style = template.styles_by_uid[ev.uniqueid]
        relay_size = style.relay_count or 4
        club_id = club_ids[cnorm]
        if tevent is None:
            issues.warn("relay_skipped",
                f"{clubs[cnorm]} relay UID {ev.uniqueid} gender={ev.gender} "
                f"({ev.age_code}) — no matching event in template")
            continue
        event_id = tevent.swim_event_id

        for club_squad_idx, squad in enumerate(squads, start=1):
            # Skip incomplete squads
            if len(squad) < relay_size:
                continue

            # Skip if any member lacks an athlete_id (e.g. not inserted)
            missing = [akey for akey, _ in squad[:relay_size]
                       if akey not in athlete_ids]
            if missing:
                relay_name = "/".join(athletes[k].last
                                      for k, _ in squad[:relay_size])
                names = ", ".join(f"{athletes[k].first} {athletes[k].last}"
                                  for k in missing)
                issues.warn("relay_skipped",
                    f"relay '{relay_name}' ({clubs[cnorm]}, UID {ev.uniqueid} "
                    f"{ev.age_code}) skipped — member not inserted: {names}")
                continue

            # Route by age-sum for Masters, by bracket label otherwise
            age_sum = None
            oldest_age = None
            if ev.age_code == "MASTERS":
                ages = [age_at(athletes[akey].birthdate)
                         for akey, _ in squad[:relay_size]]
                if any(a is None for a in ages):
                    relay_name = "/".join(athletes[akey].last
                                          for akey, _ in squad[:relay_size])
                    no_dob_names = [f"{athletes[akey].first} {athletes[akey].last}"
                                    for akey, _ in squad[:relay_size]
                                    if athletes[akey].birthdate is None]
                    issues.warn("relay_skipped",
                        f"relay '{relay_name}' ({clubs[cnorm]}, UID {ev.uniqueid} "
                        f"MASTERS) skipped — missing DOB: "
                        f"{', '.join(no_dob_names)}")
                    stats["masters_skipped_no_dob"] += 1
                    continue
                age_sum = sum(ages)
                oldest_age = max(ages)
            ag = pick_agegroup_for_relay(tevent, ev.age_code, age_sum,
                                         oldest_age=oldest_age)
            if ag is None:
                issues.warn("relay_skipped",
                    f"{clubs[cnorm]} relay UID {ev.uniqueid} "
                    f"({ev.age_code}) couldn't find an AGEGROUP "
                    f"(age_sum={age_sum})")
                continue

            _ce_key = (club_id, event_id)
            _relay_squad_counter.setdefault(_ce_key, 0)
            _relay_squad_counter[_ce_key] += 1
            stable_key = (club_id, event_id, _relay_squad_counter[_ce_key])
            if stable_key in existing_relays_stable:
                rid = existing_relays_stable[stable_key]
                for leg_no, (akey, _bt) in enumerate(squad[:relay_size],
                                                     start=1):
                    if (rid, leg_no) in existing_relay_pos:
                        continue
                    db.insert("RELAYPOSITION", {
                        "RELAYID":       rid,
                        "ATHLETEID":     athlete_ids[akey],
                        "RELAYNUMBER":   leg_no,
                        "RESULTSTATUS":  0,
                        "QTCOURSE":      0,
                        "QTISLAP":       "F",
                        "QTTIME":        INT_MAX,
                        "QTTIMING":      0,
                        "REACTIONTIME":  -32768,
                    })
                    existing_relay_pos.add((rid, leg_no))
                    stats["relayposition_new"] += 1
                continue

            next_team_no += 1
            rid = db.next_id()
            entry_time = None
            if all(bt is not None for _, bt in squad[:relay_size]):
                entry_time = sum(bt for _, bt in squad[:relay_size])
            db.insert("RELAY", {
                "RELAYID":      rid,
                "CLUBID":       club_id,
                "SWIMEVENTID":  event_id,
                "AGEGROUPID":   ag.agegroup_id,
                "GENDER":       ev.gender,
                "TEAMNUMBER":   next_team_no,
                "RELAYCODE":    next_team_no,
                "AGEMIN":       ag.amin if ag.amin is not None else 0,
                "AGEMAX":       ag.amax if ag.amax is not None else 99,
                "AGETOTAL":     age_sum if age_sum is not None else 0,
                "ATHLETES":     relay_size,
                "ENTRYTIME":    entry_time,
                "ENTRYCOURSE":  0,
                "RESULTSTATUS": 0,
                "NAME":         truncate("/".join(
                    athletes[akey].last for akey, _ in squad[:relay_size]), 100),
                "BONUSENTRY":   "F",
                "DSQNOTIFIED":  "F",
                "FINALFIX":     "F",
                "LATEENTRY":    "F",
                "NOADVANCE":    "F",
                "BACKUPTIME1":  None, "BACKUPTIME2": None, "BACKUPTIME3": None,
                "FINISHJUDGE":  None,
                "PADTIME":      None,
                "QTCOURSE":     0,
                "QTTIME":       INT_MAX,
                "QTTIMING":     0,
                "REACTIONTIME": -32768,
                "USETIMETYPE":  0,
            })
            stats["relay_new"] += 1
            existing_relays_stable[stable_key] = rid
            for leg_no, (akey, _bt) in enumerate(squad[:relay_size],
                                                 start=1):
                db.insert("RELAYPOSITION", {
                    "RELAYID":      rid,
                    "ATHLETEID":    athlete_ids[akey],
                    "RELAYNUMBER":  leg_no,
                    "RESULTSTATUS": 0,
                    "QTCOURSE":     0,
                    "QTISLAP":      "F",
                    "QTTIME":       INT_MAX,
                    "QTTIMING":     0,
                    "REACTIONTIME": -32768,
                })
                existing_relay_pos.add((rid, leg_no))
                stats["relayposition_new"] += 1

    # ----- Summary of changes -----
    print("\n" + "=" * 60)
    print("  Summary of changes")
    print("=" * 60)
    def line(label, n):
        if n:
            print(f"  +{n:<5d} {label}")
    line("new clubs",                     stats["club_new"])
    line("new athletes",                  stats["athlete_new"])
    line("athlete gender corrections",    stats["athlete_gender_fix"])
    line("athlete license fills",         stats["athlete_license_fix"])
    line("athlete birthdate fills",       stats["athlete_birthdate_fix"])
    line("athlete club changes",          stats["athlete_club_fix"])
    line("new individual entries",        stats["entry_new"])
    line("entries updated (faster time)", stats["entry_time_faster"])
    line("new relay squads",              stats["relay_new"])
    line("new relay positions",           stats["relayposition_new"])
    line("Masters entries skipped (no DOB)", stats["masters_skipped_no_dob"])
    total_changes = sum(stats.values())
    if total_changes == 0:
        print("  (no changes — database already in sync with xlsx)")
    print("=" * 60)
    print(f"\nAllocated UIDs {db._start_uid+1}..{db._uid}  "
          f"({db._uid - db._start_uid} new rows)")

    # Emit the issues section (data-quality findings) if any were collected.
    if args.issues_out:
        with open(args.issues_out, "w", encoding="utf-8") as fh:
            # Write the xlsx filename at the top of the output file so the
            # recipient knows which workbook it's from.
            fh.write(f"Data-quality report for: {args.xlsx}\n")
            fh.write(f"Generated: {dt.datetime.now():%Y-%m-%d %H:%M:%S}\n")
            issues.report("Issues found while parsing",
                          out_file=fh, full=True)
        print(f"\n(issues written to {args.issues_out})")
    else:
        issues.report("Issues found while parsing / loading",
                      full=args.issues_full)

    db.commit()
    db.close()
    if args.dry_run:
        print("\nDRY RUN complete — no changes written to disk.")
    else:
        print("\nDone. Open the .mdb in SPLASH Meet Manager to verify.")


if __name__ == "__main__":
    main()
