#!/usr/bin/env python3
"""
load_to_lenex.py
Build a Lenex 3.0 meet file (.lef XML, optionally zipped into .lxf) from
the "Attendees" sheet of an Excel registration workbook.

The resulting file can be imported into SPLASH Meet Manager (11+) or any
other Lenex-compatible tool.  It contains:

    <MEET>
      <AGEDATE>
      <POOL>
      <SESSIONS>
        <SESSION>              <!-- a single placeholder session -->
          <EVENTS>              <!-- one EVENT per distinct race ticket -->
            <SWIMSTYLE>
            <AGEGROUPS>
      <CLUBS>
        <CLUB>*                 <!-- one per attending club -->
          <ATHLETES>
            <ATHLETE>*
              <ENTRIES>
                <ENTRY>*        <!-- individual inscriptions -->
          <RELAYS>
            <RELAY>*            <!-- one per (age group, relay style) -->
              <RELAYPOSITIONS>
                <RELAYPOSITION>*

Non-race tickets (Banquet, Coach, Officiel, Cosmodôme, Couloir, Sheraton,
Priorité) are ignored.

Usage:
    python3 load_to_lenex.py --xlsx CPLC2026FINAL.xlsx --out candien.lef
    python3 load_to_lenex.py --xlsx CPLC2026FINAL.xlsx --out candien.lxf --zip
"""
from __future__ import annotations

import argparse
import datetime as dt
import os
import re
import sys
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from xml.dom import minidom
from xml.etree import ElementTree as ET

import openpyxl

# --------------------------------------------------------------------------- #
# Configuration (edit if you want different meet metadata)
# --------------------------------------------------------------------------- #
MEET_NAME   = "Canadien"
MEET_CITY   = "Laval"
MEET_NATION = "CAN"
MEET_COURSE = "LCM"
MEET_START  = dt.date(2026, 6, 19)   # day 1 — user-adjustable below
MEET_END    = dt.date(2026, 6, 21)   # day 3
AGE_DATE    = dt.date(2026, 6, 20)
PLACEHOLDER_SESSION_NAME = "Placeholder – reorganise in Meet Manager"

# --------------------------------------------------------------------------- #
# Société de Sauvetage lifesaving catalog.
#
# Same catalog as load_to_mdb.py (extracted from 30-Deux 25 octobre 2025.mdb).
# On the Lenex side we use stroke="UNKNOWN" (the standard Lenex fallback for
# non-swim events) and carry the identity in SWIMSTYLE/@name and @code.  The
# French name goes into @name (what SPLASH uses) and the English translation
# into @name2 for downstream bilingual readers.
# --------------------------------------------------------------------------- #
NON_RACE_PREFIXES = (
    "Banquet", "Coach", "Cosmod", "Couloir", "Officiel", "Priorit",
    "Sheraton",
)

LENEX_STROKE = "UNKNOWN"

# key (label, is_relay) -> (UNIQUEID, CODE, distance_m, relay_count, name_fr, name_en)
LIFESAVING_CATALOG: dict[tuple, tuple] = {
    ("Corde",             False): (504, "ID504",  12, 1,
                                   "12 m Lancer de la corde",
                                   "12 m Line Throw"),
    ("Obstacle",          False): (501, "ID501", 200, 1,
                                   "200 m Nage avec obstacles",
                                   "200 m Obstacle Swim"),
    ("Obstacle100",       False): (552, "ID552", 100, 1,
                                   "100 m Nage avec obstacles",
                                   "100 m Obstacle Swim"),
    ("Portage",           False): (502, "ID502", 100, 1,
                                   "100 m Portage du mannequin plein avec palmes",
                                   "100 m Manikin Carry w/ Fins"),
    ("Portage50",         False): (507, "ID507",  50, 1,
                                   "50 m Portage du mannequin plein",
                                   "50 m Manikin Carry"),
    ("Remorquage",        False): (506, "ID506", 100, 1,
                                   "100 m Remorquage du mannequin ½ plein + palmes",
                                   "100 m Manikin Tow with Fins"),
    ("Sauveteur d'acier", False): (508, "ID508", 200, 1,
                                   "200 m Sauveteur d'acier",
                                   "200 m Super Lifesaver"),
    ("Medley",            False): (550, "ID550", 200, 1,
                                   "200 m Medley de sauvetage",
                                   "200 m Rescue Medley"),
    ("Medley",            True):  (538, "ID538",  50, 4,
                                   "4 x 50 m Relais Medley",
                                   "4 x 50 m Rescue Medley Relay"),
    ("Obstacle",          True):  (540, "ID540",  50, 4,
                                   "4 x 50 m Relais obstacles",
                                   "4 x 50 m Obstacle Relay"),
    ("Portage",           True):  (551, "ID551",  50, 4,
                                   "4 x 50 m Relais portage du mannequin",
                                   "4 x 50 m Manikin Carry Relay"),
}

AGE_GROUPS = {
    # code -> (agemin, agemax, display)
    "1518":    (15, 18, "15-18 ans"),
    "MASTERS": (30, -1, "Maîtres"),
    "OPEN":    (-1, -1, "Toutes catégories"),
}


@dataclass(frozen=True)
class EventKey:
    age_code: str
    gender: str         # "M", "F", "X" (Lenex values)
    uniqueid: int       # SS catalog UID
    code: str           # 'ID501' etc.
    distance: int
    relay_count: int    # 1 for individual, 4 for relay
    name_fr: str
    name_en: str


def parse_ticket(ticket: str) -> EventKey | None:
    t = ticket.strip()
    for p in NON_RACE_PREFIXES:
        if t.startswith(p):
            return None
    m = re.match(r"^(15-18|MA|Open)\s+(.*)$", t)
    if not m:
        return None
    age_code = {"15-18": "1518", "MA": "MASTERS", "Open": "OPEN"}[m.group(1)]
    rest = m.group(2).strip()
    mr = re.match(r"^Relais Mixte\s+(\S+)", rest)
    if mr:
        style = mr.group(1).strip()
        cat = LIFESAVING_CATALOG.get((style, True))
        if cat is None:
            return None
        uid, code, dist, rc, fr, en = cat
        return EventKey(age_code, "X", uid, code, dist, rc, fr, en)
    mi = re.match(r"^([FM])\s+(.*)$", rest)
    if not mi:
        return None
    gender = mi.group(1)
    body = mi.group(2).strip()
    mb = re.match(r"^(.*?)(?:\s+(\d+)\s*m)?$", body)
    label = mb.group(1).strip()
    dist_txt = mb.group(2)

    # Disambiguate Portage 50 / 100 and Obstacle 100 (Masters) / 200
    lookup_label = label
    if label == "Portage" and dist_txt == "50":
        lookup_label = "Portage50"
    elif label == "Obstacle" and dist_txt == "100":
        lookup_label = "Obstacle100"

    cat = LIFESAVING_CATALOG.get((lookup_label, False))
    if cat is None:
        return None
    uid, code, dist, rc, fr, en = cat
    return EventKey(age_code, gender, uid, code, dist, rc, fr, en)


# --------------------------------------------------------------------------- #
# Time / helpers
# --------------------------------------------------------------------------- #
_TIME_RE_H = re.compile(r"^\s*(\d+):(\d{1,2}):(\d{1,2})(?:[.,](\d{1,3}))?\s*$")
_TIME_RE_M = re.compile(r"^\s*(\d+):(\d{1,2})(?:[.,](\d{1,3}))?\s*$")
_TIME_RE_S = re.compile(r"^\s*(\d+)(?:[.,](\d{1,3}))?\s*$")


def parse_best_time_ms(val: Any) -> int | None:
    """Parse a swim best time into MILLISECONDS, or None.

    Internally we normalise to ms for consistency with load_to_mdb.py
    (SPLASH's SWIMRESULT.ENTRYTIME is milliseconds); the Lenex wire
    format uses hundredths and is emitted via `ms_to_lenex()`.
    """
    if val is None:
        return None
    if isinstance(val, dt.time):
        return ((val.hour * 3600 + val.minute * 60 + val.second) * 1000
                + val.microsecond // 1000)
    if isinstance(val, dt.timedelta):
        return int(round(val.total_seconds() * 1000))
    if isinstance(val, (int, float)):
        x = float(val)
        if 0 < x < 1:
            return int(round(x * 24 * 3600 * 1000))
        if x > 0:
            return int(round(x * 1000))
        return None
    s = str(val).strip()
    if not s or s.lower() in ("nt", "n/a", "na", "-"):
        return None
    def _ms(fs): return int(((fs or "0") + "000")[:3])
    for regex, builder in [
        (_TIME_RE_H, lambda g: (int(g[0])*3600 + int(g[1])*60 + int(g[2]))*1000 + _ms(g[3])),
        (_TIME_RE_M, lambda g: (int(g[0])*60 + int(g[1]))*1000 + _ms(g[2])),
        (_TIME_RE_S, lambda g: int(g[0])*1000 + _ms(g[1])),
    ]:
        m = regex.match(s)
        if m:
            total = builder(m.groups())
            return total if total > 0 else None
    return None


def ms_to_lenex(ms: int | None) -> str | None:
    """Lenex wants swim times as 'HH:MM:SS.cc' (hundredths)."""
    if ms is None:
        return None
    total_s, sub = divmod(int(ms), 1000)
    hundredths = sub // 10
    hh, rest = divmod(total_s, 3600)
    mm, ss = divmod(rest, 60)
    return f"{hh:02d}:{mm:02d}:{ss:02d}.{hundredths:02d}"


def parse_birthdate(val: Any) -> dt.date | None:
    if val is None or (isinstance(val, float) and val != val):
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def norm_key(*parts: Any) -> str:
    s = " ".join((str(p) if p is not None else "") for p in parts).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s)


def short_code(name: str, length: int = 10) -> str:
    caps = "".join(c for c in name if c.isupper() or c.isdigit())
    if 2 <= len(caps) <= length:
        return caps
    cleaned = re.sub(r"[^A-Za-z0-9]", "", unicodedata.normalize("NFKD", name)
                     .encode("ascii", "ignore").decode("ascii"))
    return cleaned[:length].upper() or "CLUB"


# --------------------------------------------------------------------------- #
# Fuzzy duplicate detection (mirrors load_to_mdb.py)
# --------------------------------------------------------------------------- #
import difflib as _difflib
FUZZY_CLUB_THRESHOLD    = 0.90
FUZZY_ATHLETE_THRESHOLD = 0.90


def fuzzy_key(s: str) -> str:
    if s is None: return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def similarity(a: str, b: str) -> float:
    return _difflib.SequenceMatcher(None, a, b).ratio()


def find_fuzzy_club_duplicates(
        club_counts: dict[str, int],
        threshold: float = FUZZY_CLUB_THRESHOLD):
    names = sorted(club_counts.keys(), key=str.lower)
    out = []
    for i in range(len(names)):
        a = names[i]; ka = fuzzy_key(a)
        for j in range(i + 1, len(names)):
            b = names[j]; kb = fuzzy_key(b)
            if not ka or not kb:
                continue
            if ka == kb:
                out.append((a, b, 1.0, club_counts[a], club_counts[b]))
                continue
            if abs(len(ka) - len(kb)) > max(4, min(len(ka), len(kb)) // 2):
                continue
            s = similarity(ka, kb)
            if s >= threshold:
                out.append((a, b, s, club_counts[a], club_counts[b]))
    return out


def find_fuzzy_athlete_duplicates(
        inscriptions: list["Inscription"],
        threshold: float = FUZZY_ATHLETE_THRESHOLD):
    # Dedup inscriptions down to athletes (first+last+license) with the
    # first seen entry as representative.
    athletes: dict[tuple, "Inscription"] = {}
    for ins in inscriptions:
        k = (norm_key(ins.first, ins.last), (ins.license or "").strip())
        athletes.setdefault(k, ins)

    results = {
        "same_license":           [],
        "same_club_fuzzy":        [],
        "cross_club_same_person": [],
    }

    # 1) same license, different name
    by_license: dict[str, list] = {}
    for k, ins in athletes.items():
        lic = (ins.license or "").strip()
        if lic:
            by_license.setdefault(lic, []).append(ins)
    for lic, group in by_license.items():
        if len(group) < 2: continue
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                a, b = group[i], group[j]
                if fuzzy_key(f"{a.first} {a.last}") \
                        != fuzzy_key(f"{b.first} {b.last}"):
                    results["same_license"].append(
                        (f"{a.first} {a.last}", a.club,
                         f"{b.first} {b.last}", b.club, lic))

    # 2) same club, similar names
    by_club: dict[str, list] = {}
    for k, ins in athletes.items():
        by_club.setdefault(norm_key(ins.club), []).append(ins)
    for cnorm, group in by_club.items():
        if len(group) < 2: continue
        keys = [fuzzy_key(f"{ins.first} {ins.last}") for ins in group]
        for i in range(len(group)):
            ka = keys[i]
            if not ka: continue
            for j in range(i + 1, len(group)):
                kb = keys[j]
                if not kb: continue
                if ka == kb:
                    a, b = group[i], group[j]
                    if a.license != b.license:
                        results["same_club_fuzzy"].append(
                            (f"{a.first} {a.last}", a.license or "-",
                             f"{b.first} {b.last}", b.license or "-",
                             a.club, 1.0))
                    continue
                if abs(len(ka) - len(kb)) > max(4, min(len(ka), len(kb)) // 2):
                    continue
                s = similarity(ka, kb)
                if s >= threshold:
                    a, b = group[i], group[j]
                    s_first = similarity(fuzzy_key(a.first), fuzzy_key(b.first))
                    s_last  = similarity(fuzzy_key(a.last),  fuzzy_key(b.last))
                    if s_first >= 0.70 and s_last >= 0.70:
                        results["same_club_fuzzy"].append(
                            (f"{a.first} {a.last}", a.license or "-",
                             f"{b.first} {b.last}", b.license or "-",
                             a.club, s))

    # 3) cross-club same person: same normalised name + same DOB
    by_name_dob: dict[tuple, list] = {}
    for k, ins in athletes.items():
        bd = ins.birthdate
        if bd is None:
            continue
        fk = fuzzy_key(f"{ins.first} {ins.last}")
        if not fk:
            continue
        by_name_dob.setdefault((fk, bd), []).append(ins)
    for (name_k, bd_k), group in by_name_dob.items():
        clubs_ = {norm_key(ins.club) for ins in group}
        if len(clubs_) < 2: continue
        for i in range(len(group)):
            for j in range(i + 1, len(group)):
                a, b = group[i], group[j]
                if norm_key(a.club) == norm_key(b.club):
                    continue
                results["cross_club_same_person"].append(
                    (f"{a.first} {a.last}", a.club, b.club,
                     bd_k.isoformat() if bd_k else ""))

    return results


# --------------------------------------------------------------------------- #
# Read workbook
# --------------------------------------------------------------------------- #
@dataclass
class Inscription:
    first: str
    last: str
    club: str
    birthdate: dt.date | None
    license: str | None
    best_ms: int | None
    event: EventKey


# --------------------------------------------------------------------------- #
# Issue collector — data-quality warnings surfaced at end of run
# --------------------------------------------------------------------------- #
@dataclass
class Issue:
    severity: str
    category: str
    message: str
    row: int | None = None


class IssueCollector:
    def __init__(self, max_per_category: int = 10):
        self.issues: list[Issue] = []
        self.max_per_category = max_per_category

    def warn(self, category: str, message: str, row: int | None = None):
        self.issues.append(Issue("WARNING", category, message, row))

    def note(self, category: str, message: str, row: int | None = None):
        self.issues.append(Issue("NOTE", category, message, row))

    def by_category(self):
        from collections import defaultdict
        out: dict[tuple, list[Issue]] = defaultdict(list)
        for i in self.issues:
            out[(i.severity, i.category)].append(i)
        return out

    def report(self, title: str = "Issues") -> None:
        if not self.issues:
            return
        buckets = self.by_category()
        ordered = sorted(buckets.items(),
                         key=lambda kv: (kv[0][0] != "WARNING", -len(kv[1])))
        print("\n" + "=" * 60)
        print(f"  {title}")
        print("=" * 60)
        for (sev, cat), items in ordered:
            print(f"  [{sev}] {cat}: {len(items)}")
            for it in items[: self.max_per_category]:
                suffix = f" (row {it.row})" if it.row else ""
                print(f"       - {it.message}{suffix}")
            if len(items) > self.max_per_category:
                print(f"       … and {len(items) - self.max_per_category} more")
        print("=" * 60)


def read_attendees(xlsx: Path,
                   issues: IssueCollector | None = None) -> list[Inscription]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    if "Attendees" not in wb.sheetnames:
        raise SystemExit(f"Sheet 'Attendees' not found in {xlsx}")
    ws = wb["Attendees"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c or "").strip() for c in rows[0]]
    def col(n): return header.index(n)
    i_first, i_last = col("First Name"), col("Last Name")
    i_ticket = col("Ticket Type")
    i_best = col("Best time")
    i_club = col("Club")
    i_dob = col("DD/MM/YYYY")
    i_lic = col("NRAN")

    out = []
    seen_pairs: dict[tuple, int] = {}

    for row_idx, r in enumerate(rows[1:], start=2):
        if not r:
            continue
        first = r[i_first]; last = r[i_last]
        if not first or not last:
            if any(cell not in (None, "") for cell in r) and issues:
                issues.warn("missing_name",
                            "row missing first or last name", row=row_idx)
            continue
        ticket = str(r[i_ticket] or "").strip()
        if not ticket:
            if issues:
                issues.warn("missing_ticket",
                            f"{first} {last}: empty Ticket Type", row=row_idx)
            continue
        ev = parse_ticket(ticket)
        if ev is None:
            if not any(ticket.startswith(p) for p in NON_RACE_PREFIXES) \
                    and issues:
                issues.warn("unknown_ticket",
                            f"{first} {last}: unrecognised ticket {ticket!r}",
                            row=row_idx)
            continue

        raw_time = r[i_best]
        bt = parse_best_time_ms(raw_time)
        if raw_time not in (None, "") and bt is None \
                and str(raw_time).strip().lower() not in ("nt","n/a","na","-"):
            if issues:
                issues.warn("bad_time",
                            f"{first} {last} {ticket!r}: "
                            f"can't parse time {raw_time!r}",
                            row=row_idx)
        raw_dob = r[i_dob]
        bd = parse_birthdate(raw_dob)
        if raw_dob not in (None, "") and bd is None and issues:
            issues.warn("bad_birthdate",
                        f"{first} {last}: can't parse DOB {raw_dob!r}",
                        row=row_idx)

        if len(str(first)) > 30 and issues:
            issues.note("truncated_name",
                        f"first name truncated (>30 chars): {first!r}",
                        row=row_idx)
        if len(str(last)) > 50 and issues:
            issues.note("truncated_name",
                        f"last name truncated (>50 chars): {last!r}",
                        row=row_idx)

        pair_key = (norm_key(first, last), r[i_lic] or "", ev)
        if pair_key in seen_pairs and issues:
            issues.note("duplicate_entry",
                        f"{first} {last} entered in {ticket!r} again "
                        f"(first seen row {seen_pairs[pair_key]}); "
                        f"keeping best time",
                        row=row_idx)
        seen_pairs[pair_key] = row_idx

        out.append(Inscription(
            first=str(first).strip(),
            last=str(last).strip(),
            club=str(r[i_club] or "Unattached").strip(),
            birthdate=bd,
            license=(str(r[i_lic]).strip() if r[i_lic] else None),
            best_ms=bt,
            event=ev,
        ))
    return out


# --------------------------------------------------------------------------- #
# Lenex XML builder
# --------------------------------------------------------------------------- #
def lenex_gender(g: str) -> str:
    """Lenex EVENT/ATHLETE gender values: M / F / X / A."""
    return {"M": "M", "F": "F", "X": "X"}.get(g, "A")


def build_lenex(inscriptions: list[Inscription]) -> ET.ElementTree:
    # Aggregate
    clubs: dict[str, str] = {}
    athletes: dict[tuple, Inscription] = {}
    events: dict[EventKey, int] = {}  # EventKey -> event number
    ind_entries: list[tuple] = []
    relay_groups: dict[tuple, list[tuple]] = defaultdict(list)

    for ins in inscriptions:
        cnorm = norm_key(ins.club)
        clubs.setdefault(cnorm, ins.club)
        akey = (norm_key(ins.first, ins.last), ins.license or "")
        if akey not in athletes:
            athletes[akey] = ins
        events.setdefault(ins.event, None)  # placeholder, numbered below
        if ins.event.relay_count == 1:
            ind_entries.append((cnorm, akey, ins.event, ins.best_ms))
        else:
            relay_groups[(cnorm, ins.event)].append((akey, ins.best_ms))

    # Assign event numbers deterministically.
    # Sort: by catalog UID (all events of the same style together), then by
    # age bracket (15-18 < Masters < Open), then by gender with F before M.
    _AGE_ORDER = {"1518": 0, "MASTERS": 1, "OPEN": 2}
    _GENDER_ORDER = {"F": 0, "M": 1, "X": 2}   # Lenex gender strings
    sorted_events = sorted(
        events.keys(),
        key=lambda e: (e.uniqueid,
                       _AGE_ORDER.get(e.age_code, 99),
                       _GENDER_ORDER.get(e.gender, 99)))
    for i, ev in enumerate(sorted_events, start=1):
        events[ev] = i

    # Derive athlete gender (ignore relays which are mixed)
    athlete_gender: dict[tuple, str] = {}
    for ins in inscriptions:
        akey = (norm_key(ins.first, ins.last), ins.license or "")
        if ins.event.relay_count == 1 and akey not in athlete_gender:
            athlete_gender[akey] = ins.event.gender
    for akey in athletes:
        athlete_gender.setdefault(akey, "A")

    # ----- Build XML tree -----
    root = ET.Element("LENEX", {
        "version": "3.0",
        "revisiondate": "2024-12-02",
        "created": dt.datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
    })
    ctor = ET.SubElement(root, "CONSTRUCTOR", {
        "name": "load_to_lenex.py",
        "registration": "Société de Sauvetage",
        "version": "1.0",
    })
    ET.SubElement(ctor, "CONTACT", {
        "name": "Meet registration import",
        "email": "",
    })

    meets = ET.SubElement(root, "MEETS")
    meet = ET.SubElement(meets, "MEET", {
        "name": MEET_NAME,
        "city": MEET_CITY,
        "nation": MEET_NATION,
        "course": MEET_COURSE,
        "startmethod": "1",
        "timing": "AUTOMATIC",
        "touchpad": "ONESIDE",
        "masters": "F",
        "reservecount": "2",
    })
    ET.SubElement(meet, "AGEDATE",
                  {"value": AGE_DATE.isoformat(), "type": "YEAR"})
    ET.SubElement(meet, "POOL", {"lanemin": "1", "lanemax": "8"})
    ET.SubElement(meet, "FACILITY",
                  {"city": MEET_CITY, "nation": MEET_NATION})

    # ---- Sessions / Events ----
    sessions = ET.SubElement(meet, "SESSIONS")
    session = ET.SubElement(sessions, "SESSION", {
        "number": "1",
        "name": PLACEHOLDER_SESSION_NAME,
        "date": MEET_START.isoformat(),
        "daytime": "09:00",
        "course": MEET_COURSE,
    })
    evts_xml = ET.SubElement(session, "EVENTS")
    for ev in sorted_events:
        enum = events[ev]
        event_el = ET.SubElement(evts_xml, "EVENT", {
            "eventid": str(100 + enum),
            "number": str(enum),
            "order": str(enum),
            "gender": lenex_gender(ev.gender),
            "round": "TIM",
        })
        ss_attrs = {
            "distance":   str(ev.distance),
            "relaycount": str(ev.relay_count),
            "stroke":     LENEX_STROKE,
            "technique":  "NONE",
            "code":       ev.code,
            "name":       ev.name_fr,   # French — primary, consumed by SPLASH
            "name2":      ev.name_en,   # English translation for downstream readers
        }
        ET.SubElement(event_el, "SWIMSTYLE", ss_attrs)
        agemin, agemax, agename = AGE_GROUPS[ev.age_code]
        agroups = ET.SubElement(event_el, "AGEGROUPS")
        ag_attrs = {
            "agegroupid": str(1000 + enum),
            "name": agename,
            "gender": lenex_gender(ev.gender),
        }
        if agemin > 0: ag_attrs["agemin"] = str(agemin)
        else:          ag_attrs["agemin"] = "-1"
        if agemax > 0: ag_attrs["agemax"] = str(agemax)
        else:          ag_attrs["agemax"] = "-1"
        ET.SubElement(agroups, "AGEGROUP", ag_attrs)

    # ---- Clubs / Athletes / Entries / Relays ----
    clubs_xml = ET.SubElement(meet, "CLUBS")
    athlete_ids: dict[tuple, int] = {}
    athlete_uid = 10000
    for cnorm, cname in sorted(clubs.items(), key=lambda kv: kv[1].lower()):
        club_el = ET.SubElement(clubs_xml, "CLUB", {
            "name": cname,
            "shortname": short_code(cname, 30),
            "code": short_code(cname, 10),
            "nation": MEET_NATION,
        })

        # Athletes belonging to this club
        club_athletes = [
            (akey, a) for akey, a in athletes.items()
            if norm_key(a.club) == cnorm]
        if club_athletes:
            ath_xml = ET.SubElement(club_el, "ATHLETES")
            for akey, ins in sorted(club_athletes,
                                    key=lambda kv: (kv[1].last.lower(),
                                                    kv[1].first.lower())):
                athlete_uid += 1
                athlete_ids[akey] = athlete_uid
                attrs = {
                    "athleteid": str(athlete_uid),
                    "firstname": ins.first,
                    "lastname":  ins.last,
                    "gender":    lenex_gender(athlete_gender[akey]),
                }
                if ins.birthdate:
                    attrs["birthdate"] = ins.birthdate.isoformat()
                if ins.license:
                    attrs["license"] = ins.license
                ath_el = ET.SubElement(ath_xml, "ATHLETE", attrs)

                # Individual entries for this athlete
                my_entries = [
                    (ev, bt) for (ccn, ak, ev, bt) in ind_entries
                    if ak == akey]
                if my_entries:
                    entries_xml = ET.SubElement(ath_el, "ENTRIES")
                    # Deduplicate (keep best time)
                    best: dict[EventKey, int | None] = {}
                    for ev, bt in my_entries:
                        cur = best.get(ev)
                        if cur is None or (bt is not None
                                           and (cur is None or bt < cur)):
                            best[ev] = bt
                    for ev, bt in best.items():
                        attrs = {
                            "eventid": str(100 + events[ev]),
                            "status":  "",
                        }
                        et_str = ms_to_lenex(bt)
                        if et_str:
                            attrs["entrytime"] = et_str
                            attrs["entrycourse"] = MEET_COURSE
                        ET.SubElement(entries_xml, "ENTRY", attrs)

        # Relays for this club
        club_relays = [(ev, members) for (cn, ev), members in relay_groups.items()
                       if cn == cnorm]
        if club_relays:
            relays_xml = ET.SubElement(club_el, "RELAYS")
            relay_uid = 20000
            for ev, members in sorted(
                    club_relays, key=lambda x: (x[0].age_code, x[0].uniqueid)):
                # Chunk members into squads of ev.relay_count
                chunks: list[list[tuple]] = []
                buf: list[tuple] = []
                for m in members:
                    buf.append(m)
                    if len(buf) == ev.relay_count:
                        chunks.append(buf); buf = []
                if buf:
                    chunks.append(buf)   # leftover squad (may be incomplete)
                for team_no, squad in enumerate(chunks, start=1):
                    relay_uid += 1
                    entry_cs = None
                    if all(bt is not None for _, bt in squad) \
                            and len(squad) >= ev.relay_count:
                        entry_cs = sum(bt for _, bt in squad[:ev.relay_count])
                    rel_attrs = {
                        "number": str(team_no),
                        "name": f"{cname} {team_no}",
                        "agemin": str(AGE_GROUPS[ev.age_code][0]) or "-1",
                        "agemax": str(AGE_GROUPS[ev.age_code][1]) or "-1",
                        "gender": lenex_gender(ev.gender),
                    }
                    rel_el = ET.SubElement(relays_xml, "RELAY", rel_attrs)
                    ents = ET.SubElement(rel_el, "ENTRIES")
                    entry_attrs = {
                        "eventid": str(100 + events[ev]),
                    }
                    if entry_cs is not None:
                        entry_attrs["entrytime"] = ms_to_lenex(entry_cs)
                        entry_attrs["entrycourse"] = MEET_COURSE
                    entry_el = ET.SubElement(ents, "ENTRY", entry_attrs)
                    positions = ET.SubElement(entry_el, "RELAYPOSITIONS")
                    for leg, (akey, bt) in enumerate(
                            squad[:ev.relay_count], start=1):
                        if akey not in athlete_ids:
                            continue
                        ET.SubElement(positions, "RELAYPOSITION", {
                            "number": str(leg),
                            "athleteid": str(athlete_ids[akey]),
                        })

    return ET.ElementTree(root)


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", required=True, type=Path)
    ap.add_argument("--out", required=True, type=Path,
                    help="Output .lef (XML) or .lxf (zipped) path.")
    ap.add_argument("--zip", action="store_true",
                    help="Write as .lxf (ZIP) instead of raw XML.")
    ap.add_argument("--pretty", action="store_true", default=True,
                    help="Pretty-print the XML output (default).")
    args = ap.parse_args()

    if not args.xlsx.exists():
        sys.exit(f"xlsx not found: {args.xlsx}")

    print(f"Reading {args.xlsx}...")
    issues = IssueCollector()
    ins = read_attendees(args.xlsx, issues)
    print(f"  {len(ins)} race inscriptions extracted")

    # Quick stats
    clubs = {norm_key(i.club) for i in ins}
    athletes = {(norm_key(i.first, i.last), i.license or "") for i in ins}
    ev_keys = {i.event for i in ins}
    n_ind = sum(1 for e in ev_keys if e.relay_count == 1)
    n_rel = sum(1 for e in ev_keys if e.relay_count > 1)
    print(f"  clubs:    {len(clubs)}")
    print(f"  athletes: {len(athletes)}")
    print(f"  events:   {len(ev_keys)}  (individual: {n_ind}, relay: {n_rel})")
    print(f"  individual entries: "
          f"{sum(1 for i in ins if i.event.relay_count == 1)}")
    print(f"  relay entries:      "
          f"{sum(1 for i in ins if i.event.relay_count > 1)}")

    # ----- Cross-row data-quality checks -----
    # athletes with no birthdate
    ath_info: dict[tuple, Inscription] = {}
    for i in ins:
        k = (norm_key(i.first, i.last), i.license or "")
        ath_info.setdefault(k, i)
    for k, i in ath_info.items():
        if i.birthdate is None:
            issues.warn("no_dob",
                        f"{i.first} {i.last} ({i.club}) has no birthdate")
    # age-bracket mismatch on individual tickets
    bracket_by_code = {"1518": (15, 18), "MASTERS": (30, 99), "OPEN": (0, 99)}
    for i in ins:
        if i.event.relay_count != 1:
            continue
        if i.birthdate is None:
            continue
        ref = AGE_DATE
        bd = i.birthdate
        years = ref.year - bd.year
        if (ref.month, ref.day) < (bd.month, bd.day):
            years -= 1
        amin, amax = bracket_by_code[i.event.age_code]
        if years < amin or years > amax:
            issues.warn(
                "age_bracket_mismatch",
                f"{i.first} {i.last} age {years} outside ticket bracket "
                f"{i.event.age_code} ({amin}-{amax}) for {i.event.name_fr}")
    # incomplete relays
    from collections import defaultdict as _dd
    relay_members: dict[tuple, list] = _dd(list)
    for i in ins:
        if i.event.relay_count > 1:
            relay_members[(norm_key(i.club), i.event)].append(i)
    for (cnorm, ev), members in relay_members.items():
        club_disp = members[0].club
        leftovers = len(members) % ev.relay_count
        if len(members) < ev.relay_count:
            issues.warn(
                "incomplete_relay",
                f"{club_disp}: {len(members)}/{ev.relay_count} athletes "
                f"for {ev.name_fr} ({ev.age_code})")
        elif leftovers:
            n_squads = (len(members) + ev.relay_count - 1) // ev.relay_count
            issues.note(
                "extra_relay_members",
                f"{club_disp}: {len(members)} athletes for {ev.name_fr} "
                f"({ev.age_code}) — split into {n_squads} squads, the "
                f"last one has only {leftovers}/{ev.relay_count}")

    # Non-race-only clubs / athletes (informational)
    wb_all = openpyxl.load_workbook(args.xlsx, data_only=True)
    ws_all = wb_all["Attendees"]
    rows_all = list(ws_all.iter_rows(values_only=True))
    hdr = [str(c or "").strip() for c in rows_all[0]]
    i_f  = hdr.index("First Name")
    i_l  = hdr.index("Last Name")
    i_cl = hdr.index("Club")
    i_lc = hdr.index("NRAN")
    all_clubs: set[str] = set()
    all_names: set[str] = set()
    for r in rows_all[1:]:
        if not r or not r[i_f] or not r[i_l]:
            continue
        all_clubs.add(norm_key(r[i_cl] or "Unattached"))
        all_names.add(norm_key(r[i_f], r[i_l]))
    race_names = {akey[0] for akey in athletes}
    n_club_skipped = len(all_clubs - clubs)
    n_ath_skipped = len(all_names - race_names)
    if n_club_skipped:
        issues.note("non_race_only_club",
                    f"{n_club_skipped} club(s) appear only on non-race "
                    f"tickets (Banquet/Coach/Officiel/…) — not imported")
    if n_ath_skipped:
        issues.note("non_race_only_athlete",
                    f"{n_ath_skipped} attendee(s) only bought non-race "
                    f"tickets (supporters, coaches, officials, hotel) "
                    f"— not imported as athletes")

    # ----- Fuzzy duplicate detection -----
    from collections import Counter as _Counter
    club_row_counts: dict[str, int] = _Counter()
    for i in ins:
        club_row_counts[i.club] += 1
    for a, b, sim, ca, cb in find_fuzzy_club_duplicates(dict(club_row_counts)):
        issues.warn("possible_duplicate_club",
                    f"{a!r} ({ca} rows) vs {b!r} ({cb} rows) "
                    f"— similarity {sim:.2f}")
    fuzzy = find_fuzzy_athlete_duplicates(ins)
    for (name_a, club_a, name_b, club_b, lic) in fuzzy["same_license"]:
        issues.warn("license_name_mismatch",
                    f"license {lic!r}: {name_a!r} ({club_a}) vs "
                    f"{name_b!r} ({club_b}) — same license, different "
                    f"name spelling")
    for (name_a, lic_a, name_b, lic_b, club, sim) in fuzzy["same_club_fuzzy"]:
        issues.warn("possible_duplicate_athlete",
                    f"{club}: {name_a!r} (NRAN {lic_a}) vs {name_b!r} "
                    f"(NRAN {lic_b}) — similarity {sim:.2f}")
    for (name, club_a, club_b, dob) in fuzzy["cross_club_same_person"]:
        issues.warn("same_person_different_club",
                    f"{name!r} born {dob} appears in both {club_a!r} "
                    f"and {club_b!r} — probably the same person")

    tree = build_lenex(ins)

    # Serialize
    raw = ET.tostring(tree.getroot(), encoding="utf-8", xml_declaration=True)
    if args.pretty:
        raw = minidom.parseString(raw).toprettyxml(indent="  ",
                                                   encoding="utf-8")

    if args.zip or args.out.suffix.lower() == ".lxf":
        lef_name = args.out.with_suffix(".lef").name
        with zipfile.ZipFile(args.out, "w",
                             compression=zipfile.ZIP_DEFLATED) as z:
            z.writestr(lef_name, raw)
        print(f"\nWrote Lenex archive: {args.out}")
    else:
        args.out.write_bytes(raw)
        print(f"\nWrote Lenex XML: {args.out}")

    # ----- Summary of what got written + issues -----
    print("\n" + "=" * 60)
    print("  Summary")
    print("=" * 60)
    print(f"  {len(clubs):>5d}  clubs")
    print(f"  {len(athletes):>5d}  athletes")
    print(f"  {len(ev_keys):>5d}  events "
          f"({n_ind} individual, {n_rel} relay)")
    print(f"  {sum(1 for i in ins if i.event.relay_count == 1):>5d}  "
          f"individual entries")
    print(f"  {sum(len(m) for m in relay_members.values()):>5d}  "
          f"relay member entries")
    print("=" * 60)

    issues.report("Issues found while parsing")


if __name__ == "__main__":
    main()
