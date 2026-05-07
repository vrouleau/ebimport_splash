#!/usr/bin/env python3
"""
generate_test_xlsx.py
Build a deterministic, synthetic Attendees workbook that exercises every
parser path and edge case in load_to_mdb.py.

Run:
    python tests/generate_test_xlsx.py --out tests/test_attendees.xlsx

What the generated sheet contains (all numbers are deterministic — the
RNG is seeded with 20260504):

  - 100 athletes spread across 5 clubs of varied size
  - a mix of 15-18 / Open / Masters ages, both genders
  - every individual style × age-bracket × gender combination registered
    for at least one athlete (7 styles × 3 ages × 2 genders = 42 combos)
  - relay participation in all 3 relay styles × all 3 ages
  - non-race tickets (Banquet, Coach, Officiel, Cosmodome, Couloir,
    Sheraton, Priorite) for a handful of people — tests the skip path
  - deliberate data-quality defects, each flagged once:
      * 1 unknown ticket
      * 1 missing-name row
      * 1 unparseable best time
      * 1 unparseable birthdate
      * 1 duplicate (athlete, event) row
      * 2 athletes with no DOB
      * 2 athletes outside the age bracket of their ticket
      * 1 incomplete relay squad (under-subscribed)
      * 1 relay squad with leftover members

A parallel `tests/EXPECTED_ISSUES.md` documents exactly what each script
should report — use it as the manual regression reference.
"""
from __future__ import annotations

import argparse
import datetime as dt
import random
from collections import defaultdict
from pathlib import Path

import openpyxl

SEED = 20260504
AGE_DATE = dt.date(2026, 6, 20)   # same as the loaders

# Five fictitious clubs of varied size
CLUBS = [
    "Aurora Test Club",         # largest (30 athletes)
    "Béluga Sauvetage",         # accents in name, 25 athletes
    "Cedar Creek LSC",          # 20 athletes
    "Dauphins de l'Est",        # 15 athletes
    "Elite Rescue",             # smallest (10 athletes)
]
CLUB_SIZES = [30, 25, 20, 15, 10]   # sum=100

# Sample first/last names so athletes are recognisable
FIRST_F = ["Alice","Béatrice","Chloé","Diane","Emma","Frédérique","Gabrielle",
           "Héloïse","Inès","Juliette","Karine","Léa","Maude","Noémie","Océane",
           "Pénélope","Quinn","Rosalie","Sophie","Tania","Ursula","Valérie",
           "Willow","Xiomara","Yasmine","Zoé"]
FIRST_M = ["Alexandre","Benoît","Christophe","David","Émile","François","Gabriel",
           "Hugo","Isaac","Jérôme","Kevin","Liam","Mathis","Nathaniel","Olivier",
           "Philippe","Quentin","Raphaël","Samuel","Thomas","Ulysse","Vincent",
           "William","Xavier","Yannick","Zachary"]
LAST = ["Tremblay","Gagnon","Roy","Côté","Bouchard","Gauthier","Morin","Lavoie",
        "Fortin","Gagné","Ouellet","Pelletier","Bélanger","Lévesque","Bergeron",
        "Leblanc","Paquette","Girard","Simard","Boucher","Caron","Beaulieu",
        "Cloutier","Dubois","Poirier","Fournier","Lapointe","Leclerc","Lemieux",
        "Mercier"]

INDIVIDUAL_TICKETS_OPEN = [
    "Open F Corde",              "Open M Corde",
    "Open F Medley",             "Open M Medley",
    "Open F Obstacle",           "Open M Obstacle",
    "Open F Portage 100m",       "Open M Portage 100m",
    "Open F Portage 50m",        "Open M Portage 50m",
    "Open F Remorquage",         "Open M Remorquage",
    "Open F Sauveteur d'acier",  "Open M Sauveteur d'acier",
]
INDIVIDUAL_TICKETS_1518 = [t.replace("Open ", "15-18 ") for t in INDIVIDUAL_TICKETS_OPEN]
# Obstacle uses the 100 m distance for Masters
INDIVIDUAL_TICKETS_MA = [
    "MA F Corde",                "MA M Corde",
    "MA F Medley",               "MA M Medley",
    "MA F Obstacle 100 m",       "MA M Obstacle 100 m",
    "MA F Portage 100m",         "MA M Portage 100m",
    "MA F Portage 50m",          "MA M Portage 50m",
    "MA F Remorquage",           "MA M Remorquage",
    "MA F Sauveteur d'acier",    "MA M Sauveteur d'acier",
]
RELAY_TICKETS = {
    "1518":    ["15-18 Relais Mixte Medley", "15-18 Relais Mixte Obstacle",
                "15-18 Relais Mixte Portage"],
    "OPEN":    ["Open Relais Mixte Medley",  "Open Relais Mixte Obstacle",
                "Open Relais Mixte Portage"],
    "MASTERS": ["MA Relais Mixte Medley",    "MA Relais Mixte Obstacle",
                "MA Relais Mixte Portage"],
}
NON_RACE_TICKETS = [
    "Banquet",
    "Banquet Officiel 3 jours",
    "Coach",
    "Cosmodôme 2 nuits",
    "Couloir de nage 17h-18h",
    "Officiel 3 jours",
    "Priorité - SERC",
    "Sheraton 3 nuits",
]

HEADER = ["First Name", "Last Name", "Email", "Ticket Type", "Best time",
          "Club", "DD/MM/YYYY", "NRAN", "Teammate(s) + NRAN", "Visitor(s)",
          "Allergies", "Day(s) present", "Favorite song"]


def random_time(rng: random.Random, rough_seconds: float) -> str:
    """Return an 'mm:ss.cc' string around `rough_seconds`, ±20%."""
    secs = rough_seconds * (0.8 + 0.4 * rng.random())
    mm   = int(secs // 60)
    ss   = int(secs % 60)
    cs   = int((secs - int(secs)) * 100)
    return f"{mm}:{ss:02d}.{cs:02d}"


def dob_for_age(target_age: int, rng: random.Random) -> str:
    """Return a DOB such that `age_at(dob, AGE_DATE)` == target_age."""
    # Pick a birthday somewhere not on June-20 so the age is unambiguous.
    month = rng.randint(1, 12)
    day   = rng.randint(1, 28)
    birth_year = AGE_DATE.year - target_age
    if (month, day) > (AGE_DATE.month, AGE_DATE.day):
        birth_year -= 1
    return f"{day:02d}/{month:02d}/{birth_year}"


def license_code(first: str, last: str, seq: int) -> str:
    """Short NRAN-style license: first 2 of last + first 2 of first + seq."""
    def strip(s):
        import unicodedata as u
        s = u.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
        return "".join(c for c in s if c.isalpha()).upper()
    return f"{strip(last)[:2]}{strip(first)[:2]}{seq:02d}"


def build_athletes(rng: random.Random):
    """Return a list of athlete dicts.  100 athletes total.

    Each has:
      first, last, email, club, nran, dob (str dd/mm/yyyy),
      gender ('F'|'M'), age_code ('1518'|'OPEN'|'MASTERS'), age (int).
    """
    athletes = []
    seq = 0
    # To guarantee we can build an over-subscribed relay for Béluga Open and
    # a full squad for Dauphins Masters, we plant a minimum number of
    # athletes of the right bracket in those two clubs before filling the
    # rest randomly.
    guaranteed: dict[str, list[str]] = {
        "Béluga Sauvetage":  ["OPEN"] * 5,     # 5 Open athletes (relay+leftover)
        "Dauphins de l'Est": ["MASTERS"] * 4,  # 4 Masters athletes (full squad)
        "Cedar Creek LSC":   ["1518"] * 2,     # 2 under-subscribed relay athletes
        "Aurora Test Club":  ["1518"] * 4 + ["OPEN"] * 4,
        "Elite Rescue":      ["1518"] * 4,
    }
    for club, size in zip(CLUBS, CLUB_SIZES):
        planted = 0
        planned = guaranteed.get(club, [])
        for j in range(size):
            seq += 1
            gender = rng.choice(["F", "M"])
            # Plant the required age code for the first N athletes of this club
            if j < len(planned):
                code = planned[j]
                if code == "1518":
                    age = rng.randint(15, 18)
                elif code == "OPEN":
                    age = rng.randint(19, 29)
                else:
                    age = rng.randint(30, 60)
                age_code = code
            else:
                bucket = rng.random()
                if bucket < 0.40:
                    age_code = "1518"
                    age = rng.randint(15, 18)
                elif bucket < 0.80:
                    age_code = "OPEN"
                    age = rng.randint(19, 29)
                else:
                    age_code = "MASTERS"
                    age = rng.randint(30, 60)
            first = rng.choice(FIRST_F if gender == "F" else FIRST_M)
            last  = rng.choice(LAST)
            athletes.append({
                "first": first, "last": last,
                "email": f"test.{seq:03d}@example.com",
                "club":  club,
                "nran":  license_code(first, last, seq),
                "dob":   dob_for_age(age, rng),
                "gender": gender,
                "age_code": age_code,
                "age": age,
            })
    return athletes


def pick_individual_tickets(ath: dict, rng: random.Random) -> list[str]:
    """Pick 2-5 individual tickets appropriate for the athlete's age+gender."""
    pool = {
        "1518":    INDIVIDUAL_TICKETS_1518,
        "OPEN":    INDIVIDUAL_TICKETS_OPEN,
        "MASTERS": INDIVIDUAL_TICKETS_MA,
    }[ath["age_code"]]
    # keep only tickets matching the athlete's gender.  Every ticket has the
    # pattern "<age prefix> <F|M> <rest>" — peek at the token after the prefix.
    def ticket_gender(t: str) -> str:
        for prefix in ("15-18 ", "Open ", "MA "):
            if t.startswith(prefix):
                return t[len(prefix):len(prefix) + 1]
        return "?"
    my_pool = [t for t in pool if ticket_gender(t) == ath["gender"]]
    if not my_pool:
        return []
    n = rng.randint(2, min(5, len(my_pool)))
    return rng.sample(my_pool, n)


def rough_seconds_for_ticket(ticket: str) -> float:
    """Rough expected time (in seconds) for a ticket — used to seed realistic
    best times."""
    t = ticket.lower()
    if "corde" in t:       return 25
    if "medley" in t:      return 180
    if "obstacle 100" in t:return 85
    if "obstacle" in t:    return 165
    if "portage 50" in t:  return 45
    if "portage" in t:     return 95
    if "remorquage" in t:  return 90
    if "sauveteur" in t:   return 190
    return 120


def write_rows(ws, rows):
    for r in rows:
        ws.append(r + [None] * (len(HEADER) - len(r)))


def build_rows(athletes, rng):
    """Build the full list of Attendees rows including deliberate defects."""
    rows = []

    # 1) Normal individual tickets
    # Track Corde entries per club to pair them up later
    corde_by_club: dict[str, list[dict]] = defaultdict(list)
    for ath in athletes:
        for ticket in pick_individual_tickets(ath, rng):
            rows.append([
                ath["first"], ath["last"], ath["email"], ticket,
                random_time(rng, rough_seconds_for_ticket(ticket)),
                ath["club"], ath["dob"], ath["nran"], None,
            ])
            if "Corde" in ticket:
                corde_by_club[ath["club"]].append(ath)

    # Pair up Corde athletes (duo relay) — add teammate field
    for club, corde_athletes in corde_by_club.items():
        pairs = list(zip(corde_athletes[::2], corde_athletes[1::2]))
        for a, b in pairs:
            # Find their Corde rows and add teammate
            for row in rows:
                if row[0] == a["first"] and row[1] == a["last"] and "Corde" in (row[3] or ""):
                    row[8] = f"{b['first']} {b['last']} {b['nran']}"
                elif row[0] == b["first"] and row[1] == b["last"] and "Corde" in (row[3] or ""):
                    row[8] = f"{a['first']} {a['last']} {a['nran']}"

    # 2) Relay tickets — sprinkle mixed relays per club per age bracket.
    # Aim for most to have >=4 people (full squads); also build one
    # under-subscribed squad and one with leftover members.
    def add_relay_rows(club, age_code, size):
        """Pick `size` athletes from this club/age bracket (any gender)
        and register them for all 3 relay styles of that bracket."""
        pool = [a for a in athletes
                if a["club"] == club and a["age_code"] == age_code]
        chosen = rng.sample(pool, min(size, len(pool)))
        for ath in chosen:
            # List teammates (all other chosen members), newline-separated
            teammates = "\n".join(
                f"{t['first']} {t['last']} {t['nran']}"
                for t in chosen if t is not ath
            )
            for rt in RELAY_TICKETS[age_code]:
                rows.append([
                    ath["first"], ath["last"], ath["email"], rt,
                    random_time(rng, rough_seconds_for_ticket(rt)),
                    ath["club"], ath["dob"], ath["nran"], teammates,
                ])

    # Club 1: 4 athletes for 15-18 relay (complete squad)
    add_relay_rows("Aurora Test Club",   "1518",    4)
    # Club 1: 4 athletes for Open relay
    add_relay_rows("Aurora Test Club",   "OPEN",    4)
    # Club 2: 5 athletes for Open relay (1 leftover — "extra_relay_members")
    add_relay_rows("Béluga Sauvetage",   "OPEN",    5)
    # Club 3: only 2 athletes for 15-18 relay (INCOMPLETE)
    add_relay_rows("Cedar Creek LSC",    "1518",    2)
    # Club 4: 4 athletes for Masters relay
    add_relay_rows("Dauphins de l'Est",  "MASTERS", 4)
    # Club 5: 4 athletes for 15-18 relay
    add_relay_rows("Elite Rescue",       "1518",    4)

    # 3) Non-race tickets — coverage for the skip path
    for i, ticket in enumerate(NON_RACE_TICKETS):
        ath = athletes[i]
        rows.append([
            ath["first"], ath["last"], ath["email"], ticket, None,
            ath["club"], ath["dob"], ath["nran"],
        ])

    # 4) Deliberate data-quality defects (each flagged once)

    # 4a) Unknown ticket
    rows.append([
        "Zach", "Unknown", "z@x.com", "Not A Real Ticket",
        "1:00.00", "Aurora Test Club", "01/01/1995", "UNKNOWN01",
    ])
    # 4b) Missing first name (but has other fields)
    rows.append([
        None, "NoFirstName", "nofn@x.com", "Open F Obstacle",
        "1:00.00", "Aurora Test Club", "01/01/1995", "NOFIRST01",
    ])
    # 4c) Bad best time
    rows.append([
        "Bob", "BadTime", "bt@x.com", "Open M Obstacle",
        "not-a-time", "Aurora Test Club", "01/01/1995", "BADTIME01",
    ])
    # 4d) Bad birthdate
    rows.append([
        "Beth", "BadDOB", "bd@x.com", "Open F Obstacle",
        "1:50.00", "Aurora Test Club", "maybe 2001?", "BADDOB01",
    ])
    # 4e) Duplicate (athlete, event) — same pair twice; second row has slower time
    dup_ath = athletes[0]    # re-use an existing athlete
    rows.append([
        dup_ath["first"], dup_ath["last"], dup_ath["email"], "Open F Obstacle",
        "1:35.00", dup_ath["club"], dup_ath["dob"], dup_ath["nran"],
    ])
    rows.append([
        dup_ath["first"], dup_ath["last"], dup_ath["email"], "Open F Obstacle",
        "1:42.00", dup_ath["club"], dup_ath["dob"], dup_ath["nran"],
    ])
    # 4f) Two athletes with no DOB
    rows.append([
        "Nora", "NoDOB", "ndob@x.com", "Open F Remorquage",
        "1:40.00", "Aurora Test Club", None, "NODOB01",
    ])
    rows.append([
        "Noel", "NoDOB", "ndob2@x.com", "Open M Remorquage",
        "1:30.00", "Béluga Sauvetage", None, "NODOB02",
    ])
    # 4g) Two athletes with age outside their bracket
    #   Under-age: DOB implies age 13, registered in 15-18
    under_age_dob = dob_for_age(13, rng)
    rows.append([
        "Under", "AgeTooYoung", "under@x.com", "15-18 F Obstacle",
        "2:00.00", "Cedar Creek LSC", under_age_dob, "UNDERAGE1",
    ])
    #   Over-age: DOB implies age 20, registered in 15-18
    over_age_dob = dob_for_age(20, rng)
    rows.append([
        "Over", "AgeTooOld", "over@x.com", "15-18 M Obstacle",
        "1:45.00", "Cedar Creek LSC", over_age_dob, "OVERAGE1",
    ])

    # 4h) Fuzzy-duplicate scenarios (each should trigger one warning)

    # 4h-i) Club spelling variant — "Béluga Sauvetage" vs "Beluga Sauvetage"
    #       (accent-folded identical) and "Aurora Test Club" vs
    #       "Aurora Test  Club" (double-space).
    rows.append([
        "Varianta", "Clubtyperow", "clubtv1@x.com", "Open F Obstacle",
        "1:33.00", "Beluga Sauvetage",           # no accent
        "01/01/1998", "CLUBVAR01",
    ])
    rows.append([
        "Variantb", "Clubtyperow", "clubtv2@x.com", "Open M Obstacle",
        "1:28.00", "Aurora Test  Club",          # double space
        "01/01/1998", "CLUBVAR02",
    ])

    # 4h-ii) Same license, different name spelling — "Henri Chiu" vs
    #        "Henri Tsz Hin Chiu" (both use the same NRAN).
    rows.append([
        "Henri", "Chiu", "chiu1@x.com", "Open M Remorquage",
        "1:34.00", "Aurora Test Club", "15/05/1999", "CHIU_SAME",
    ])
    rows.append([
        "Henri Tsz Hin", "Chiu", "chiu2@x.com", "Open M Remorquage",
        "1:34.00", "Aurora Test Club", "15/05/1999", "CHIU_SAME",
    ])

    # 4h-iii) Athlete name typo within same club — "Stephen Kennedy" vs
    #         "Stphen Kennedy" (missing 'e').
    rows.append([
        "Stephen", "Kennedy", "k1@x.com", "Open M Obstacle",
        "2:22.00", "Elite Rescue", "06/02/1971", "KENN01",
    ])
    rows.append([
        "Stphen", "Kennedy", "k2@x.com", "Open M Obstacle",
        "2:23.00", "Elite Rescue", "06/02/1971", "KENN02",
    ])

    # 4h-iv) Cross-club same-person — same name+DOB in two clubs.
    rows.append([
        "Gabrielle", "Fortin", "gf1@x.com", "Open F Obstacle",
        "1:40.00", "Dauphins de l'Est", "03/03/2000", "GFORT_A",
    ])
    rows.append([
        "Gabrielle", "Fortin", "gf2@x.com", "Open F Obstacle",
        "1:41.00", "Elite Rescue", "03/03/2000", "GFORT_B",
    ])

    return rows


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", type=Path,
                    default=Path(__file__).parent / "test_attendees.xlsx")
    args = ap.parse_args()

    rng = random.Random(SEED)
    athletes = build_athletes(rng)
    # Sanity: make sure every (age_code × gender × style) combo appears at
    # least once.  Because athletes are random we may miss a combo; add
    # make-up tickets at the end if so.

    rows = build_rows(athletes, rng)

    # Coverage check — add make-up rows to hit every individual combo
    needed = (INDIVIDUAL_TICKETS_OPEN + INDIVIDUAL_TICKETS_1518
              + INDIVIDUAL_TICKETS_MA)
    covered = {r[3] for r in rows if r[3]}
    missing = [t for t in needed if t not in covered]

    def ticket_prefix(t: str) -> str:
        for p in ("15-18 ", "Open ", "MA "):
            if t.startswith(p):
                return p
        return ""

    def ticket_gender(t: str) -> str:
        p = ticket_prefix(t)
        return t[len(p):len(p) + 1] if p else "?"

    age_code_of_prefix = {"15-18 ": "1518", "Open ": "OPEN", "MA ": "MASTERS"}

    for ticket in missing:
        p = ticket_prefix(ticket)
        needed_code   = age_code_of_prefix.get(p, "")
        needed_gender = ticket_gender(ticket)
        for a in athletes:
            if a["age_code"] == needed_code and a["gender"] == needed_gender:
                rows.append([
                    a["first"], a["last"], a["email"], ticket,
                    random_time(rng, rough_seconds_for_ticket(ticket)),
                    a["club"], a["dob"], a["nran"],
                ])
                break

    # Write workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendees"
    ws.append(HEADER)
    write_rows(ws, rows)

    # Add the other two sheets that the real workbook has (empty placeholders
    # so loader behaviour is identical)
    wb.create_sheet("Athlètes")
    wb.create_sheet("Autres")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"Wrote {args.out}")
    print(f"  athletes: {len(athletes)}")
    print(f"  data rows: {len(rows)}  (excluding header)")


if __name__ == "__main__":
    main()
